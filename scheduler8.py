from copy import deepcopy
import random
import os
import sys
import itertools
import openpyxl
import get_info
from stack import Slot_Stack

# sys.setrecursionlimit(10 ** 6)


def getWorkBook():
    return openpyxl.load_workbook("sheets/Book.xlsx")


def getWorkBookLab():
    return openpyxl.load_workbook("sheets/BookLab.xlsx")


def getTT():
    TT = [['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', ''],
          ['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', ''],
          ['', '', '', '', 'Lunch', '', '', '']]
    return TT


# days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
day_cell = ['7', '8', '9', '10', '11']
time_cell = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

# sections
sections = get_info.get_sections_and_mentors(1)
sec_list = sections.keys()

# initializing workbooks
workbooks = {
    'Lib': getWorkBook()
}

for sec in sec_list:
    workbooks[sec] = getWorkBook()

# initializing sections and mentors
mentors = sections

# getting faculty information
faculty = get_info.getFacultyInfo()
# fac_list = []

for fac in faculty:
    workbooks[fac.getFName()] = getWorkBook()
    # fac_list.append(fac.getFName())

# getting labs
labs = get_info.get_labs()
# print(labs)
for lab in labs.keys():
    workbooks[lab] = getWorkBookLab()

# Dictionary containing the name of the sec, labs and faculty with their respective worksheets
section_sheets = {}
lab_sheets = {}
faculty_sheets = {}

for sec in sections.keys():
    section_sheets[sec] = workbooks[sec].active

section_sheets['Lib'] = workbooks['Lib'].active

for lab in labs.keys():
    lab_sheets[lab] = workbooks[lab].active

for fac in faculty:
    faculty_sheets[fac.getFName()] = workbooks[fac.getFName()].active


# setting load
def set_load():
    global av_fac
    av_fac = set()
    load_list = get_info.getFacLoad()
    for fac in faculty:
        fac.resetFLoad()

    while load_list:
        for l in load_list:
            for fac in faculty:
                if fac.getFId() == l[0]:
                    l.remove(l[0])
                    if 'CS3' in l[0]:
                        sec = 'CSE 3' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                        av_fac.add(fac)

                        break
                    elif 'CS5' in l[0]:
                        sec = 'CSE 5' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                    elif 'AL3' in l[0]:
                        sec = 'AIML 3' + l[1]
                        fac.setFLoad(sec, l)
                        # print("Added")
                        load_list.remove(l)
                        av_fac.add(fac)
                        break
                    elif 'AL5' in l[0]:
                        sec = 'AIML 5' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                    elif 'CY3' in l[0]:
                        sec = 'CY 3' + l[1]
                        fac.setFLoad(sec, l)
                        # print("Added")
                        load_list.remove(l)
                        av_fac.add(fac)
                        break
                    elif 'CY5' in l[0]:
                        sec = 'CY 5' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)


set_load()


def reset_load():
    global load_list
    load_list = get_info.getFacLoad()


# Dictionary containing the name of the sec, lab and faculty with their timetable structure
section_tt = {}
lab_tt = {}
faculty_tt = {}

for sec in sections.keys():
    section_tt[sec] = getTT()

section_tt['Lib'] = getTT()

for lab in labs.keys():
    lab_tt[lab] = getTT()

for fac in faculty:
    faculty_tt[fac.getFName()] = getTT()

all_tt = {
    'Faculty': faculty_tt,
    'Labs': lab_tt,
    'Sections': section_tt
}

all_presets = {
    'Faculty': [],
    'Labs': [],
    'Sections': []
}


def reset_presets():
    for i in all_presets.keys():
        all_presets[i] = []


section_labs = {}
section_labs_fac = {}
section_th = {}
section_th_fac = {}

for sec in sections.keys():
    section_labs[sec] = {}
    section_labs_fac[sec] = {}

for sec in sections.keys():
    if '3' in sec:
        subjects = get_info.get_sub_by_section(sec)
        l = []
        section_th[sec] = []
        section_th_fac[sec] = {}
        for sub in subjects.keys():
            if subjects[sub][1] == 'B':
                l.append(sub)
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

            elif subjects[sub][1] == 'L':
                l.append(sub)

            elif subjects[sub][1] == 'T':
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

        section_labs[sec][l[0] + '/' + l[1]] = []
        section_labs[sec][l[2] + '/' + l[3]] = []
        section_labs[sec][l[4]] = []

        section_labs_fac[sec][l[0] + '/' + l[1]] = []
        section_labs_fac[sec][l[2] + '/' + l[3]] = []
        section_labs_fac[sec][l[4]] = []
    elif '5' in sec:
        subjects = get_info.get_sub_by_section(sec)
        # print(subjects)
        section_th[sec] = []
        section_th_fac[sec] = {}
        for sub in subjects.keys():
            if subjects[sub][1] == 'B':
                section_labs[sec][sub] = []
                section_labs_fac[sec][sub] = []
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

            elif subjects[sub][1] == 'L':
                section_labs[sec][sub] = []
                section_labs_fac[sec][sub] = []

            elif subjects[sub][1] == 'T':
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

# for i, j in section_th.items():
#     print(i, j)

for sec in sections.keys():
    for sub in section_labs[sec].keys():
        if '/' in sub:
            s1 = sub[0:5]
            s2 = sub[6:]
            l = ['', '']
            for fac in faculty:
                if sec in fac.getFLoad().keys():
                    load = fac.getFLoad()[sec]
                else:
                    continue
                for i in load:
                    if i[0] == s1:
                        l[0] = fac
                        break
                for i in load:
                    if i[0] == s2:
                        l[1] = fac
                        break
            section_labs_fac[sec][sub] = l
        else:
            for fac in faculty:
                if sec in fac.getFLoad().keys():
                    load = fac.getFLoad()[sec]
                else:
                    continue
                comp = 0
                for i in load:

                    if i[0] == sub:
                        section_labs_fac[sec][sub].append(fac)
                        comp = 1
                        break
                if comp == 1:
                    break

for sec in sec_list:
    for sub in section_th[sec]:
        for fac in faculty:
            if sec in fac.getFLoad().keys():
                load = fac.getFLoad()[sec]
            else:
                continue
            comp = 0
            for i in load:
                if i[0] == sub:
                    section_th_fac[sec][sub] = fac
                    comp = 1
            if comp == 1:
                break

# setting regular information in the sheets.
for sheet in section_sheets.keys():
    if sheet == 'Lib':
        section_sheets['Lib']['A4'] = 'Branch: CS/AI/CY'
        section_sheets['Lib']['C4'] = 'Sections: ALL'
        section_sheets['Lib']['F7'] = 'LUNCH'
        continue
    section_sheets[sheet]['A4'] = 'Branch: ' + sheet[0:-2]
    section_sheets[sheet]['C4'] = 'Section: ' + sheet[-2:]
    section_sheets[sheet]['F7'] = 'LUNCH'

for sheet in lab_sheets.keys():
    lab_sheets[sheet]['A4'] = 'Branch: ALL'
    lab_sheets[sheet]['C4'] = 'Section: ALL'
    lab_sheets[sheet]['F7'] = 'LUNCH'
    lab_sheets[sheet]['E4'] = 'Room No.: ' + sheet + " S"
    lab_sheets[sheet].title = sheet

for fac in faculty:
    faculty_sheets[fac.getFName()]['A4'] = 'Name: ' + fac.getFName()
    faculty_sheets[fac.getFName()]['F7'] = 'LUNCH'
    faculty_sheets[fac.getFName()]['E4'] = ''

# initializing reference index for Excel sheets
fac_info_row = ['14', '15', '16', '17', '18', '19', '20', '21']

# setting information section in the Excel sheets.
for sheet in section_sheets:
    subjects = get_info.get_sub_by_section(sheet)
    if 'Lib' in sheet:
        continue
    section_sheets[sheet]['G27'] = mentors[sheet].getFName()
    section_sheets[sheet]['I27'] = mentors[sheet].getFPhone()
    # print(subjects)
    for fac in faculty:
        if sheet in fac.getFLoad().keys():
            f_load = fac.getFLoad()[sheet]
            if '3' in sheet:
                for i in range(len(f_load)):
                    # print(fac_info_row[int(f_load[i][0][4]) - 2])
                    section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 1]] = f_load[i][0]
                    if 'CS' in f_load[i][0]:
                        section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = subjects[f_load[i][0]][0]
                    elif 'AL' in f_load[i][0]:
                        section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = subjects[f_load[i][0]][0]
                    elif 'CY' in f_load[i][0]:
                        section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = subjects[f_load[i][0]][0]
                    section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFName()
                    section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFPhone()

            elif '5' in sheet:
                for i in range(len(f_load)):
                    if '8' in f_load[i][0]:
                        section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 2]] = f_load[i][0]
                        if 'CS' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = \
                                subjects[f_load[i][0]][0]
                        elif 'AL' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = \
                                subjects[f_load[i][0]][0]
                        elif 'CY' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = \
                                subjects[f_load[i][0]][0]
                        section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 2]] = fac.getFName()
                        section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 2]] = fac.getFPhone()
                    else:
                        section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 1]] = f_load[i][0]
                        if 'CS' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = \
                                subjects[f_load[i][0]][0]
                        elif 'AL' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = \
                                subjects[f_load[i][0]][0]
                        elif 'CY' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = \
                                subjects[f_load[i][0]][0]
                        section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFName()
                        section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFPhone()

# lab allotment
lab_allotment = get_info.get_lab_allotment()
# for i,j in lab_allotment.items():
#     print(i,j)

for sec in sections.keys():
    if '3' in sec:
        for sub in section_labs[sec].keys():
            if '/' in sub:
                s1 = sub[0:5]
                s2 = sub[6:]
                li = ['', '']
                l = lab_allotment[s1]
                for i in l:
                    if sec[-1] == i[0]:
                        li[0] = i[1]
                        break
                l = lab_allotment[s2]
                for i in l:
                    if sec[-1] == i[0]:
                        li[1] = i[1]
                        break
                section_labs[sec][sub] = li
            else:
                l = lab_allotment[sub]
                for i in l:
                    if sec[-1] == i[0]:
                        section_labs[sec][sub] = [i[1]]
                        break
    elif '5' in sec:
        for sub in section_labs[sec].keys():
            l = lab_allotment[sub]
            for i in l:
                if sec[-1] == i[0]:
                    section_labs[sec][sub] = [i[1]]
                    break


section_tt_copy = deepcopy(section_tt)
lab_tt_copy = deepcopy(lab_tt)
faculty_tt_copy = deepcopy(faculty_tt)

all_tt = {
    'Faculty': faculty_tt,
    'Labs': lab_tt,
    'Sections': section_tt
}

all_presets = {
    'Faculty': [],
    'Labs': [],
    'Sections': []
}


def reset_tt():
    global all_presets, section_tt, lab_tt, faculty_tt, section_tt_copy, lab_tt_copy, faculty_tt_copy
    section_tt = deepcopy(section_tt_copy)
    lab_tt = deepcopy(lab_tt_copy)
    faculty_tt = deepcopy(faculty_tt_copy)
    presets = all_presets['Sections']
    for i in presets:
        sec = i[2]
        day = i[0]
        time = i[1]
        data = i[3]
        section_tt[sec][day][time] = data


def available_lab(sec, lab, day, time):
    global section_labs_fac, section_labs, faculty_tt, section_tt, lab_tt

    labs_fac = section_labs_fac[sec]
    labs = section_labs[sec]

    if section_tt[sec][day][time] == section_tt[sec][day][time + 1] == '':
        pass
    else:
        return False

    if lab in section_tt[sec][day]:
        return False

    if '3' in sec:
        if '07' not in lab and '08' not in lab:
            if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
                pass
            else:
                return False
            if lab_tt[labs[lab][1]][day][time] == '' and lab_tt[labs[lab][1]][day][time + 1] == '':
                pass
            else:
                return False
            if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
                pass
            else:
                return False
            if faculty_tt[labs_fac[lab][1].getFName()][day][time] == '' and \
                    faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] == '':
                pass
            else:
                return False

            if time == 0:
                fac1 = faculty_tt[labs_fac[lab][0].getFName()][day]
                fac2 = faculty_tt[labs_fac[lab][1].getFName()][day]
                if (fac1[time + 2] == '' and fac1[time + 3] != '' and fac2[time + 2] == '' and fac2[time + 3] != '') or \
                        (fac1[time + 3] == '' and fac1[time + 2] != '' and fac2[time + 3] == '' and fac2[
                            time + 2] != '') or \
                        (fac1[time + 2] == '' and fac1[time + 3] != '' and fac2[time + 3] == '' and fac2[
                            time + 2] != '') or \
                        (fac1[time + 3] == '' and fac1[time + 2] != '' and fac2[time + 2] == '' and fac2[
                            time + 3] != '') or \
                        (fac1[time + 2] == '' and fac1[time + 3] == '' and fac2[time + 2] == '' and fac2[
                            time + 3] == '') or \
                        (fac1[time + 2] == 'NA' or fac1[time + 3] == '' and fac2[time + 2] == '' and fac2[
                            time + 3] == 'NA'):
                    pass
                else:
                    return False
            elif time == 2:
                fac1 = faculty_tt[labs_fac[lab][0].getFName()][day]
                fac2 = faculty_tt[labs_fac[lab][1].getFName()][day]
                if (fac1[time - 2] == '' and fac1[time - 1] != '' and fac2[time - 2] == '' and fac2[time - 1] != '') or \
                        (fac1[time - 1] == '' and fac1[time - 2] != '' and fac2[time - 1] == '' and fac2[
                            time - 2] != '') or \
                        (fac1[time - 2] == '' and fac1[time - 1] != '' and fac2[time - 1] == '' and fac2[
                            time - 2] != '') or \
                        (fac1[time - 1] == '' and fac1[time - 2] != '' and fac2[time - 2] == '' and fac2[
                            time - 1] != '') or \
                        (fac1[time - 2] == '' and fac1[time - 1] == '' and fac2[time - 2] == '' and fac2[
                            time - 1] == '') or\
                        (fac1[time - 2] == 'NA' and fac1[time - 1] == '' and fac2[time - 2] == '' and fac2[
                            time - 1] == 'NA'):
                    pass
                else:
                    return False

            # cse_labs_fac[lab][0].getFLoad()[sec]

        else:
            if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
                pass
            else:
                return False
            if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
                pass
            else:
                return False

            if time == 0:
                if faculty_tt[labs_fac[lab][0].getFName()][day][time + 2] == '' or \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 3] == '':
                    pass
                else:
                    return False



            if time == 2:
                if faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] == '' or \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time - 1] == '' or\
                        faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] != '' or\
                        faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] != '':
                    pass
                else:
                    return False
    elif '5' in sec:
        if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
            pass
        else:
            return False
        if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
            pass
        else:
            return False

        if time == 0:
            if faculty_tt[labs_fac[lab][0].getFName()][day][time + 2] == '' or \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 3] == '':
                pass
            else:
                return False

        if time == 2:
            if faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] == '' or \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time - 1] == '':
                pass
            else:
                return False


    return True


slots = []
visited = {}


def visited_dict():
    for sec in sec_list:
        visited[sec] = False


visited_dict()


def reset_visited():
    for i in visited.keys():
        visited[i] = False


last_sec = ''

# lab_count={}
sem_3_sec = []
sem_5_sec = []
for i in sec_list:
    if '3' in i:
        sem_3_sec.append(i)
    elif '5' in i:
        sem_5_sec.append(i)
# sem_3_sec.reverse()
# sem_5_sec.reverse()
print(sem_3_sec, sem_5_sec)

lab_counter = {}


def reset_lab_counter():
    for lab in lab_tt.keys():
        lab_counter[lab] = 0


reset_lab_counter()

lab_day_counter = {}

for sec in sec_list:
    lab_day_counter[sec] = {}
    for day in [0, 1, 2, 3, 4]:
        lab_day_counter[sec][day] = 2


def reset_lab_day_counter():
    for sec in lab_day_counter.keys():
        for day in lab_day_counter[sec].keys():
            lab_day_counter[sec][day] = 2


def reset_lab_day_counter_sec(sec):
    for day in lab_day_counter[sec].keys():
        lab_day_counter[sec][day] = 2


lab_days = [0, 1, 2, 3, 4]
lab_time = [0, 2, 6]
comp_lab = len(sections) * 5
count = 0
slot_dict = {}
for sec in sections:
    slot_dict[sec] = []


def change_slot(day, time, sec, lab):
    pass


slot_stack = Slot_Stack()
end_count = 5 * len(sections)
end_counter = 0
lab_index = 0

sec_lab_slot = []


def set_labs(sec, lab_list, lab_count, lab_index, lab_days, lab_time):
    global end_counter, end_count, section_tt, section_tt, faculty_tt, lab_tt, prev_sec, prev_lab_days, prev_labs

    print(sec)
    labs_fac = section_labs_fac[sec]
    labs = section_labs[sec]

    slot_list = list(itertools.product(lab_days, lab_time))
    random.shuffle(slot_list)

    lab = random.choice(lab_list)

    for slot in slot_list:

        day = slot[0]
        time = slot[1]

        slot_list.remove((day, time))

        if available_lab(sec, lab, day, time):
            section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab + 'lab'
            if len(labs[lab]) == 2:
                lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                lab_tt[labs[lab][1]][day][time] = lab_tt[labs[lab][1]][day][time + 1] = sec
                faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub
                faculty_tt[labs_fac[lab][1].getFName()][day][time] = \
                    faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] = sub
            else:
                lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub

            lab_count[lab] -= 1
            end_counter += 1

            if end_counter == end_count:
                return True

            if lab_count[lab] == 0:
                lab_list.remove(lab)

            lab_day_counter[sec][day] -= 1

            # if lab_day_counter[sec][day] == 0:
            lab_days.remove(day)

            sec_lab_slot.append([sec, day, time, lab, lab_days, lab_list, slot_list, lab_index, lab_count])

            # print(sec_lab_slot[sec])
            if sum(lab_count.values()) == 0:
                prev_sec = sec
                lab_days = [0, 1, 2, 3, 4]
                lab_time = [0, 2, 6]
                lab_index += 1
                if lab_index < len(sections):
                    pass
                else:
                    return True
                sec = list(sections.keys())[lab_index]
                lab_list = list(section_labs[sec].keys())
                lab_count = {}
                for lab in lab_list:
                    if '/' in lab:
                        lab_count[lab] = 2
                    else:
                        lab_count[lab] = 1

            # prev_lab_fac = deepcopy(labs_fac)
            # prev_labs = deepcopy(labs)
            # print(sec, len(sec_lab_slot), sec_lab_slot[-1])
            # sec_lab_slot[sec].append([day, time, lab])
            if set_labs(sec, lab_list, lab_count, lab_index, lab_days, lab_time):
                return True

            l = sec_lab_slot.pop()
            sec, day, time, lab, lab_days, lab_list, slot_list, lab_index, lab_count = \
                l[0], l[1], l[2], l[3], l[4], l[5], l[6], l[7], l[8]
            # slot_list.remove((day,time))
            random.shuffle(slot_list)
            if lab not in lab_list:
                lab_list.append(lab)
            lab_day_counter[sec][day] += 1
            end_counter -= 1
            lab_count[lab] += 1

            if day in lab_days:
                pass
            else:
                lab_days.append(day)
            # print(labs)
            # labs_fac = section_labs_fac[sec]
            # labs = section_labs[sec]
            # print(sec, labs, lab)
            # slot_list.append((day, time))
            section_tt[sec][day][time] = section_tt[sec][day][time + 1] = ''
            if len(labs[lab]) == 2:
                lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = ''
                lab_tt[labs[lab][1]][day][time] = lab_tt[labs[lab][1]][day][time + 1] = ''
                faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = ''
                faculty_tt[labs_fac[lab][1].getFName()][day][time] = \
                    faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] = ''
            else:
                lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = ''
                faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = ''
        else:
            slot_list.append((day, time))
    return False


def call_fxn1():
    # return True
    lab_days = [0, 1, 2, 3, 4]
    lab_time = [0, 2, 6]

    # if lab_index == 14:
    #     break
    lab_index = 0
    sec = list(sections.keys())[lab_index]
    lab_list = list(section_labs[sec].keys())
    lab_count = {}
    for lab in lab_list:
        # print(lab_count)
        if '/' in lab:
            lab_count[lab] = 2
        else:
            lab_count[lab] = 1
    print(sec, lab_count)
    ans = set_labs(sec, lab_list, lab_count, lab_index, lab_days, lab_time)
    for sec in sections.keys():
        print(sec)
        for i in section_tt[sec]:
            print(i)
    if ans:
        return True


# ans = call_fxn1()


def checkEmpty(sec):
    global section_tt
    for i in section_tt[sec]:
        # print(sec)
        if '' in i:
            # tt=section_tt[sec]
            # for i in tt:
            #     print(i)
            return True
    return False


def available_theory(sec, day, time, sub):
    global section_tt, faculty_tt, section_th_fac

    th_fac = section_th_fac[sec]

    # sec1 = tt[day][0:4]
    # sec2 = tt[day][5:]

    if section_tt[sec][day][time] == '':
        pass
    else:
        return False

    # if sub in section_tt[sec][day]:
    #     return False
    if section_tt[sec][day].count(sub) <= 2:
        pass
    else:
        return False

    if sub == 'M/L':
        fac = mentors[sec]
        if section_tt['Lib'][day][time] == '':
            pass
        else:
            return False
    else:
        fac = th_fac[sub]

    # if faculty_tt[fac.getFName()][day].count('') <= 2:
    #     return False

    fac_day = faculty_tt[fac.getFName()][day]
    if fac_day[time] == '':
        pass
    else:
        return False

    if time == 0:
        if fac_day[1] == fac_day[2] == fac_day[3] == '':
            pass
        elif (fac_day[1] == '') and ((fac_day[2].__contains__('/') or fac_day[2].__contains__('lab'))):
            pass
        elif (fac_day[time + 1] == fac_day[time + 3] == '') and (fac_day[time + 2] != ''):
            pass
        elif (fac_day[time + 1] != '') and (fac_day[time + 2] == '') and (fac_day[time + 3] != ''):
            pass
        elif (fac_day[time + 1] == '') and (fac_day[time + 2] != '') and (fac_day[time + 3] != ''):
            pass
        else:
            return False

    elif time == 1:
        if fac_day[0] == fac_day[2] == fac_day[3] == '':
            pass
        elif (fac_day[time - 1] == fac_day[time + 2] == '') and (fac_day[time + 1] != ''):
            pass
        elif fac_day[0] != '' and fac_day[2] == '' and fac_day[3] == '':
            pass
        elif fac_day[0] != '' and fac_day[3] != '' and fac_day[2] == '':
            pass
        elif fac_day[0] == '' and ((fac_day[2].__contains__('/') or fac_day[2].__contains__('lab'))):
            pass
        else:
            return False

    elif time == 2:
        if fac_day[0] == fac_day[1] == fac_day[3] == '':
            pass
        elif ((fac_day[0].__contains__('/') or fac_day[0].__contains__('lab'))) and fac_day[3] == '':
            pass
        elif fac_day[0] != '' and fac_day[1] == '' and fac_day[3] != '':
            pass
        elif fac_day[0] == fac_day[1] == '' and fac_day[3] != '':
            pass
        else:
            return False

    elif time == 3:
        if fac_day[0] == fac_day[1] == fac_day[2] == '':
            pass
        elif ((fac_day[0].__contains__('/') or fac_day[0].__contains__('lab'))) and fac_day[2] == '':
            pass
        elif fac_day[0] == fac_day[1] == '' and fac_day[2] != '':
            pass
        elif fac_day[0] != '' and fac_day[1] != '' and fac_day[2] == '':
            pass
        elif fac_day[0] == '' and fac_day[2] == '' and fac_day[1] != '':
            pass
        elif fac_day[0] != '' and fac_day[2] != '' and fac_day[1] == '':
            pass
        else:
            return False

    elif time == 5:
        if fac_day[6] == '' and fac_day[7] == '':
            pass
        elif fac_day[6] == '' and fac_day[7] != '':
            pass
        elif fac_day[6] != '' and fac_day[7] == '':
            pass
        elif fac_day[6].__contains__('/') or fac_day[6].__contains__('lab'):
            pass
        else:
            return False

    elif time == 6:
        if fac_day[5] == fac_day[7] == '':
            pass
        elif fac_day[5] == '' and fac_day[7] != '':
            pass
        elif fac_day[5] != '' and fac_day[7] == '':
            pass
        else:
            return False

    else:
        if fac_day[5] == fac_day[6] == '':
            pass
        elif fac_day[5] == '' and fac_day[6] != '':
            pass
        elif fac_day[5] != '' and fac_day[6] == '':
            pass
        else:
            return False

    return True


sec_ml_set_status = {}
for sec in sections.keys():
    sec_ml_set_status[sec] = False

# sec_th_slot = []
end_count2 = len(sections.keys()) * 25 - (len(sections.keys()))
end_counter2 = 0

sec_ml_slot = []


def set_ml():
    global section_tt, faculty_tt
    for sec in sections.keys():
        while True:
            day = random.choice([0, 1, 2, 3, 4])
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            fac = mentors[sec]
            if available_theory(sec, day, time, 'M/L'):
                section_tt[sec][day][time] = 'M/L'
                section_tt['Lib'][day][time] = sec
                faculty_tt[fac.getFName()][day][time] = sec + '\n' + 'M/L'
                sec_ml_slot.append([sec, day, time, fac])
                break

# sec_th_slots={
#     'CSE 3':[],
#     'AIML 3':[],
#     'CY 3':[],
#     'CSE 5': [],
#     'AIML 5': [],
#     'CY 5': []
# }


def set_theory(sec, th_list, th_count, th_index, th_days, th_time, slot_list, th_fac, sec_list, sec_th_slot):
    global end_counter2, end_count2, section_tt, faculty_tt
    # if n checkEmpty(sec)
    # th_fac = section_th_fac[sec]

    # slot_list = list(itertools.product(th_days,th_time))
    # tt = section_tt[sec]
    # slot_list = []
    # for day in range(len(tt)):
    #     for time in range(len(tt[day])):
    #         if tt[day][time] == '':
    #             slot_list.append([day, time])

    # random.shuffle(slot_list)

    random.shuffle(slot_list)
    random.shuffle(th_list)

    print(sec, th_list)
    # sub = random.choice(th_list)
    # l1=list(itertools.product(slot_list,[sub]))
    slot = random.choice(slot_list)
    # slot_list.remove(slot)
    for sub in th_list:
        # day,time=slot[0][0],slot[0][1]
        day, time = slot[0], slot[1]

        # sub=slot[1]
        # print(slot)
        # l1.remove(slot)
        if available_theory(sec, day, time, sub):
            # slot_list.remove(slot)

            fac = th_fac[sub]
            section_tt[sec][day][time] = sub
            faculty_tt[fac.getFName()][day][time] = sub

            th_count[sub] -= 1
            end_counter2 += 1

            # if end_counter2 == end_count2:
            #     return True

            if th_count[sub] == 0:
                th_list.remove(sub)
            # slot_list.remove(slot)
            sec_th_slot.append([sec, day, time, sub, th_list, th_count, th_index, th_days, th_time, slot_list, th_fac, sec_list, sec_th_slot])
            # slot_list.remove(slot)

            if not checkEmpty(sec):
                # return True
                th_index += 1
                if th_index == len(sections.keys()):
                    return True
                sec = sec_list[th_index]
                th_list = deepcopy(section_th[sec])
                th_count = {}
                for sub in th_list:
                    if len(th_list) == 4:
                        th_count[sub] = 6
                    elif len(th_list) == 5:
                        th_count[sub] = 5
                tt = section_tt[sec]
                slot_list = []
                for day in range(len(tt)):
                    for time in range(len(tt[day])):
                        if tt[day][time] == '':
                            slot_list.append([day, time])
                random.shuffle(slot_list)

            # print(sec, sub)
            # if sec == 'CSE 3A' or sec == 'CSE 3B':
            #     pass
            # else:
            #     return True
            opt = set_theory(sec, th_list, th_count, th_index, th_days, th_time, slot_list, th_fac, sec_list, sec_th_slot)

            if opt:
                return True
            # elif opt=='rerun':
            #     print("returning")
            #     return 'rerun'

            if sec_th_slot == []:
                return False

            l = sec_th_slot.pop()
            sec, day, time, sub, th_list, th_count, th_index, th_days, th_time, slot_list, th_fac, sec_list = \
                l[0], l[1], l[2], l[3], l[4], l[5], l[6], l[7], l[8], l[9], l[10], l[11]

            if sub not in th_list:
                th_list.append(sub)
            th_count[sub] += 1
            fac = th_fac[sub]
            section_tt[sec][day][time] = ''
            faculty_tt[fac.getFName()][day][time] = ''
            # slot_list.append([day, time])
            end_counter2 -= 1

    # while sec_th_slot != []:
    #     l = sec_th_slot[sec[0:-1]].pop()
    #     sec, day, time, sub = l[0], l[1], l[2], l[3]
    #     th_fac = section_th_fac[sec]
    #     fac = th_fac[sub]
    #     section_tt[sec][day][time] = ''
    #     faculty_tt[fac.getFName()][day][time] = ''
    # return False


        # else:
        #     slot_list.append(slot)
    # return False
    # if (slot_list == []) and (end_counter2 != end_count2):
    # if (slot_list == []):
    #     print("fxn called")
    # # for sec in sections.keys():
    # #     print(sec)
    # #     for i in section_tt[sec]:
    # #         print(i)
    # # end_counter2 = 0
    #     while sec_th_slot != []:
    #         l = sec_th_slot.pop()
    #         sec, day, time, sub = l[0], l[1], l[2], l[3]
    #         th_fac = section_th_fac[sec]
    #         fac = th_fac[sub]
    #         section_tt[sec][day][time] = ''
    #         faculty_tt[fac.getFName()][day][time] = ''
    #     return False
    # # while sec_ml_slot != []:
    # #     l = sec_ml_slot.pop()
    # #     sec, day, time, fac = l[0], l[1], l[2], l[3]
    # #     section_tt[sec][day][time] = ''
    # #     section_tt['Lib'][day][time] = ''
    # #     faculty_tt[fac.getFName()][day][time] = ''
    # print("cleared all")
    # return False

    # set_ml()
    # a=call_fxn2()
    # if a:
    #     return True
    # exit()
    # return False


# sections.__reversed__()
def call_fxn2(sec_list):
    th_days = [0, 1, 2, 3, 4]
    th_time = [0, 1, 2, 3, 5, 6, 7]

    # if lab_index == 14:
    #     break
    th_index = 0
    sec = sec_list[th_index]
    th_list = deepcopy(section_th[sec])
    th_fac = section_th_fac[sec]
    sec_th_slot=[]
    th_count = {}
    for sub in th_list:
        if len(th_list) == 4:
            th_count[sub] = 6
        elif len(th_list) == 5:
            th_count[sub] = 5
    print("th count", th_count, th_list)
    tt = section_tt[sec]
    slot_list = []
    for day in range(len(tt)):
        for time in range(len(tt[day])):
            if tt[day][time] == '':
                slot_list.append([day, time])

    print(sec, th_count)

    ans = set_theory(sec, th_list, th_count, th_index, th_days, th_time, slot_list, th_fac, sec_list, sec_th_slot)
    return ans
    # while not ans:
    #     set_ml()
    #     ans = set_theory(sec, th_list, th_count, th_index, th_days, th_time, slot_list)
    #     if ans:
    #         return True
    # else:
    #     return False
    # else:
    #     call_fxn2()

    # if end_count2 != end_counter2:
    #     print("running again")
    #     set_ml()
    #     call_fxn2()

    # for sec in sections.keys():
    #     print(sec)
    #     for i in section_tt[sec]:
    #         print(i)


# ans2=False
# if ans:
#     print("labs set")
#     set_ml()
#     ans2 = call_fxn2()


def copy_tt_to_excel():
    for sec in sec_list:
        print(sec)
        tt = section_tt[sec]
        for i in tt:
            print(i)
        sheet = section_sheets[sec]
        labs = section_labs[sec]
        labs_fac = section_labs_fac[sec]
        th_fac = section_th_fac[sec]
        for day in [0, 1, 2, 3, 4]:
            for time in [0, 1, 2, 3, 5, 6, 7]:
                if tt[day][time] == '':
                    continue

                elif tt[day][time] == 'NA' or tt[day][time] == 'Na' or tt[day][time]=='TNP':
                    sheet[time_cell[time] + day_cell[day]] = tt[day][time]
                    continue

                elif tt[day][time] == 'M/L':
                    sheet[time_cell[time] + day_cell[day]] = 'M/L'
                    section_sheets['Lib'][time_cell[time] + day_cell[day]] = sec
                    faculty_sheets[mentors[sec].getFName()][time_cell[time] + day_cell[day]] = sec + '\n' + 'M/L'
                    continue

                elif tt[day][time] == sheet[time_cell[time] + day_cell[day]].value:
                    continue

                # elif 'lab' in tt[day][time]:
                elif len(tt[day][time]) == 5:
                    sub = tt[day][time]
                    fac = th_fac[sub]
                    sheet[time_cell[time] + day_cell[day]] = sub
                    # print(fac.getFName())
                    # print(sub)
                    # print(faculty_sheets[fac.getFName()][time_cell[time-1] + day_cell[day]].value)
                    # print('Time',time)
                    faculty_sheets[fac.getFName()][time_cell[time] + day_cell[day]] = sec + "\n" + sub
                    continue

                elif tt[day][time] != '' and sheet[time_cell[time] + day_cell[day]].value == None:
                    if ('/' in tt[day][time]) and (time not in [1, 3, 7]):
                        lab = tt[day][time][0:11]
                        sheet.merge_cells(time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        sheet[time_cell[time] + day_cell[day]] = lab + "\n" + \
                                                                 labs[lab][0] + "/" + labs[lab][1]
                        if '8' in lab:
                            sheet['E' + fac_info_row[int(lab[4]) - 2]] = labs[lab][0] + ' Building LNCT&S'
                            sheet['E' + fac_info_row[int(lab[10]) - 2]] = labs[lab][1] + ' Building LNCT&S'
                        else:
                            sheet['E' + fac_info_row[int(lab[4]) - 1]] = labs[lab][0] + ' Building LNCT&S'
                            sheet['E' + fac_info_row[int(lab[10]) - 1]] = labs[lab][1] + ' Building LNCT&S'
                        lab_sheets[labs[lab][0]][time_cell[time] + day_cell[day]] = \
                            lab_sheets[labs[lab][1]][time_cell[time] + day_cell[day]] = sec
                        faculty_sheets[labs_fac[lab][0].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][1].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][0].getFName()][time_cell[time] + day_cell[day]] = \
                            faculty_sheets[labs_fac[lab][1].getFName()][
                                time_cell[time] + day_cell[day]] = sec + "\n" + lab + "\n" + labs[lab][0] + '/' + \
                                                                   labs[lab][1]

                        continue
                    elif ('lab' in tt[day][time]) and (time not in [1, 3, 7]):
                        if tt[day][time][0:5] in section_labs[sec].keys():
                            if time in [0, 2, 6]:
                                pass
                            else:
                                continue
                            labs_fac = section_labs_fac[sec]
                            lab = tt[day][time][0:5]
                            sheet.merge_cells(
                                time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                            sheet[time_cell[time] + day_cell[day]] = lab + "\n" + labs[lab][0]
                            if '8' in lab:
                                sheet['E' + fac_info_row[int(lab[4]) - 2]] = labs[lab][0] + ' Building LNCT&S'
                                # sheet['E' + fac_info_row[int(lab[10]) - 2]] = labs[lab][1] + ' Building LNCT&S'
                            else:
                                sheet['E' + fac_info_row[int(lab[4]) - 1]] = labs[lab][0] + ' Building LNCT&S'
                                # sheet['E' + fac_info_row[int(lab[10]) - 1]] = labs[lab][1] + ' Building LNCT&S'
                            lab_sheets[labs[lab][0]][time_cell[time] + day_cell[day]] = sec
                            faculty_sheets[labs_fac[lab][0].getFName()].merge_cells(
                                time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                            faculty_sheets[labs_fac[lab][0].getFName()][
                                time_cell[time] + day_cell[day]] = sec + "\n" + lab + '\n' + labs[lab][0]
                            continue


def set_sheets():
    files = next(os.walk('C:/Users/Swadesh Sharma/Desktop/SchedulXpert/'))[1]
    if not files:
        n = 0
    else:
        n = len(files)
    newpath = 'C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test ' + str(n + 1)

    # Creating new directories
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        os.makedirs(newpath + '/Faculty')
        os.makedirs(newpath + '/Sections')
        os.makedirs(newpath + '/Labs')

    # saving the files in the directory creates
    for i in workbooks.keys():
        if 'Lib' in i or 'CS' in i or 'AIML' in i or 'CY' in i:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Sections/" + i + ".xlsx")
        elif 'Dr' in i or 'Prof' in i:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Faculty/" + i + ".xlsx")
        else:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Labs/" + i + ".xlsx")


# def fxn():
#     call_fxn1()
#     set_ml()
#     ans=call_fxn2()
#     if ans:
#         return True

call_fxn1()
# for sec in sections.keys():
#     c=0
#     days=[0,1,2,3,4]
#     while c!=2:
#         day=random.choice(days)
#         time=random.choice([0,2,6])
#         if section_tt[sec][day][time]==section_tt[sec][day][time+1]=='':
#             section_tt[sec][day][time] = section_tt[sec][day][time + 1] = 'TNP'
#             c+=1
#             days.remove(day)


set_ml()
# for sec in sec_list:
#     print(sec)
#     tt = section_tt[sec]
#     for i in tt:
#         print(i)

def fxn2(sec):
    # call_fxn1()
    # set_ml()
    # ans=call_fxn2(sec)
    while True:
        if call_fxn2(sec) == True:
            return True


import threading

threads = []
sec_list_type={'CSE 3':[],
               'AIML 3':[],
               'CY 3':[],
               'CSE 5': [],
               'AIML 5': [],
               'CY 5': []
               }
for i in sections.keys():
    if 'CSE 3' in i:
        sec_list_type['CSE 3'].append(i)
    elif 'AIML 3' in i:
        sec_list_type['AIML 3'].append(i)
    elif 'CY 3' in i:
        sec_list_type['CY 3'].append(i)
    elif 'CSE 5' in i:
        sec_list_type['CSE 5'].append(i)
    elif 'AIML 5' in i:
        sec_list_type['AIML 5'].append(i)
    elif 'CY 5' in i:
        sec_list_type['CY 5'].append(i)


for i in sec_list_type.values():
    # print(i)
    # if 'CSE' in i[0]:
    #     pass
    # else:
    #     continue
    th = threading.Thread(target=fxn2, args=(i,))
    threads.append(th)
    th.start()
    th.join()

for i in threads:
    i.join()

# sec_list_type={
#     'AIML 3':[],
#     'AIML 5':[]
# }
#
# for i in sections.keys():
#     # if 'CSE 3' in i:
#     #     sec_list_type['CSE 3'].append(i)
#     if 'AIML 3' in i:
#         sec_list_type['AIML 3'].append(i)
#     # elif 'CY 3' in i:
#     #     sec_list_type['CY 3'].append(i)
#     # elif 'CSE 5' in i:
#     #     sec_list_type['CSE 5'].append(i)
#     elif 'AIML 5' in i:
#         sec_list_type['AIML 5'].append(i)
#     # elif 'CY 5' in i:
#     #     sec_list_type['CY 5'].append(i)
#
#
# for i in sec_list_type.values():
#     # print(i)
#     # if 'CSE' in i[0]:
#     #     pass
#     # else:
#     #     continue
#     th = threading.Thread(target=fxn2, args=(i,))
#     threads.append(th)
#     th.start()
#     # th.join()
#
# for i in threads:
#     i.join()
#
# sec_list_type={
#     'CY 3':[],
#     'CY 5':[]
# }
#
# for i in sections.keys():
#     # if 'CSE 3' in i:
#     #     sec_list_type['CSE 3'].append(i)
#     # elif 'AIML 3' in i:
#     #     sec_list_type['AIML 3'].append(i)
#     if 'CY 3' in i:
#         sec_list_type['CY 3'].append(i)
#     # elif 'CSE 5' in i:
#     #     sec_list_type['CSE 5'].append(i)
#     # elif 'AIML 5' in i:
#     #     sec_list_type['AIML 5'].append(i)
#     elif 'CY 5' in i:
#         sec_list_type['CY 5'].append(i)
#
#
# for i in sec_list_type.values():
#     # print(i)
#     # if 'CSE' in i[0]:
#     #     pass
#     # else:
#     #     continue
#     th = threading.Thread(target=fxn2, args=(i,))
#     threads.append(th)
#     th.start()
#     # th.join()
#
# for i in threads:
#     i.join()

# threads = []
# for i in sections.keys():
#     if '5' in i:
#         pass
#     else:
#         continue
#     th = threading.Thread(target=fxn2, args=(i,))
#     threads.append(th)
#     th.start()
#
# for i in threads:
#     i.join()

# print(sec_list)
# print('Started')
# setTimeTables()
# for sec in sec_list:
#     print(sec)
#     tt = section_tt[sec]
#     for i in tt:
#         print(i)
# if fxn2():
copy_tt_to_excel()
set_sheets()
print("set")

# print('finished')
