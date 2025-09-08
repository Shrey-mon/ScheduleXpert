import os
import random

import openpyxl
from subject_info import *
from get_info import *
from copy import deepcopy


def getWorkBook():
    return openpyxl.load_workbook("sheets/Book.xlsx")


def getLabWorkBook():
    return openpyxl.load_workbook("sheets/BookLab.xlsx")


def getCsAWb():
    return openpyxl.load_workbook("sheets/Book1.xlsx")


def getCsBWb():
    return openpyxl.load_workbook("sheets/Book2.xlsx")


def getCsCWb():
    return openpyxl.load_workbook("sheets/Book3.xlsx")


def getAlAWb():
    return openpyxl.load_workbook("sheets/Book4.xlsx")


def getAlBWb():
    return openpyxl.load_workbook("sheets/Book5.xlsx")


def getAlCWb():
    return openpyxl.load_workbook("sheets/Book6.xlsx")


def getCyAWb():
    return openpyxl.load_workbook("sheets/Book7.xlsx")


def getCyBWb():
    return openpyxl.load_workbook("sheets/Book8.xlsx")


def getFacWorkBook():
    return openpyxl.load_workbook("sheets/Faculty.xlsx")


def getTT():
    TT = [['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', ''],
          ['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', ''],
          ['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', '']]
    return TT


# days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
day_cell = ['7', '8', '9', '10', '11', '12']
time_cell = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

workbooks = {
    'Lib': getWorkBook(),
    'CSE 4A': getCsAWb(),
    'CSE 4B': getCsBWb(),
    'CSE 4C': getCsCWb(),
    'AIML 4A': getAlAWb(),
    'AIML 4B': getAlBWb(),
    'AIML 4C': getAlCWb(),
    'CY 4A': getCyAWb(),
    'CY 4B': getCyBWb()
}

# List of all the sections, labs and faculty members
section_list = list(workbooks.keys())
lab_list = list(getLabInfo().keys())
fac_list = getFacultyInfo()

# Creating workbooks for Faculty Members
for fac in fac_list:
    workbooks[fac.getFName()] = getFacWorkBook()

# Creating workbooks for Labs
for lab in getLabInfo().keys():
    workbooks[lab] = getLabWorkBook()

workbooks_copy = deepcopy(workbooks)

# def reset_workbooks():
#     global workbooks
#     workbooks = deepcopy(workbooks)


# Dictionary containing the name of the sec, labs and faculty with their respective worksheets
section_sheets = {}
lab_sheets = {}
faculty_sheets = {}

for sec in section_list:
    section_sheets[sec] = workbooks[sec].active

for lab in lab_list:
    lab_sheets[lab] = workbooks[lab].active

for fac in fac_list:
    faculty_sheets[fac.getFName()] = workbooks[fac.getFName()].active

section_sheets_copy = deepcopy(section_sheets)
lab_sheets_copy = deepcopy(lab_sheets)
faculty_sheets_copy = deepcopy(faculty_sheets)

# resets the workbooks
# def resetWB():
#     global section_sheets, faculty_sheets, lab_sheets
#     section_sheets = (section_sheets_copy)
#     faculty_sheets = (faculty_sheets_copy)
#     lab_sheets = (lab_sheets_copy)


# Dictionary containing the name of the sec, lab and faculty with their timetable structure
section_tt = {}
lab_tt = {}
faculty_tt = {}

for sec in section_list:
    section_tt[sec] = getTT()

for lab in lab_list:
    lab_tt[lab] = getTT()

for fac in fac_list:
    faculty_tt[fac.getFName()] = getTT()

section_tt_copy = deepcopy(section_tt)
lab_tt_copy = deepcopy(lab_tt)
faculty_tt_copy = deepcopy(faculty_tt)

all_tt = {
    'Faculty': faculty_tt,
    'Labs': lab_tt,
    'Sections': section_tt
}

all_presets = {
    'Faculty': [],
    'Labs': [],
    'Sections': []
}


# def reset_tt():
#     global section_tt, lab_tt, faculty_tt, section_tt_copy, lab_tt_copy, faculty_tt_copy
#     section_tt = deepcopy(section_tt_copy)
#     lab_tt = deepcopy(lab_tt_copy)
#     faculty_tt = deepcopy(faculty_tt_copy)


def reset_tt():
    global all_presets, section_tt, lab_tt, faculty_tt, section_tt_copy, lab_tt_copy, faculty_tt_copy
    section_tt = deepcopy(section_tt_copy)
    lab_tt = deepcopy(lab_tt_copy)
    faculty_tt = deepcopy(faculty_tt_copy)

    for type in all_presets.keys():
        presets = all_presets[type]
        if type == 'Faculty':
            for i in presets:
                fac = i[2]
                day = i[0]
                time = i[1]
                data = i[3]
                faculty_tt[fac][day][time] = data
        elif type == 'Labs':
            for i in presets:
                lab = i[2]
                day = i[0]
                time = i[1]
                data = i[3]
                lab_tt[lab][day][time] = data
        elif type == 'Sections':
            for i in presets:
                sec = i[2]
                day = i[0]
                time = i[1]
                data = i[3]
                section_tt[sec][day][time] = data


# Getting dictionaries containing subject and lab information
cse_subjects = getSubInfoCS()
aiml_subjects = getSubInfoAL()
cy_subjects = getSubInfoCY()
lab_info = getLabInfo()
cse_sub_code_names = getCSSubNames()
aiml_sub_code_names = getALSubNames()
cy_sub_code_names = getCYSubNames()


# Copying existing data from Excel workbook to tt
def copy_existing_data():
    for sec in section_sheets.keys():
        sheet = section_sheets[sec]
        tt = section_tt[sec]
        for day in range(len(tt)):
            for time in [0, 1, 2, 3, 5, 6, 7]:
                if sheet[time_cell[time] + day_cell[day]].value != None:
                    tt[day][time] = sheet[time_cell[time] + day_cell[day]].value


copy_existing_data()

sec_sheet_copy = deepcopy(section_sheets)

# def set_sec_sheet_copy():
#     global section_sheets
#     section_sheets = sec_sheet_copy


# The list of half capacity labs and full capacity labs
half_labs = []
full_labs = []

for i in lab_list:
    if lab_info[i][0] == 'h':
        half_labs.append(i)
    else:
        full_labs.append(i)

half_labs_copy = half_labs.copy()
full_labs_copy = full_labs.copy()

# setting lab load
lab_load = {}
for i in lab_list:
    lab_load[i] = 16

# setting regular information in the sheets.
for sheet in section_sheets.keys():
    if sheet == 'Lib':
        section_sheets['Lib']['A4'] = 'Branch: CS/AI/CY'
        section_sheets['Lib']['C4'] = 'Sections: ALL'
        section_sheets['Lib']['F7'] = 'LUNCH'
        continue
    section_sheets[sheet]['A4'] = 'Branch: ' + sheet[0:-2]
    section_sheets[sheet]['C4'] = 'Section: ' + sheet[-2:]
    section_sheets[sheet]['F7'] = 'LUNCH'

for sheet in lab_sheets.keys():
    lab_sheets[sheet]['A4'] = 'Branch: ALL'
    lab_sheets[sheet]['C4'] = 'Section: ALL'
    lab_sheets[sheet]['F7'] = 'LUNCH'
    lab_sheets[sheet]['E4'] = 'Room No.: ' + sheet + " S"
    lab_sheets[sheet].title = sheet

for fac in fac_list:
    faculty_sheets[fac.getFName()]['A4'] = 'Name: ' + fac.getFName()
    faculty_sheets[fac.getFName()]['F7'] = 'LUNCH'
    faculty_sheets[fac.getFName()]['E4'] = ''

# fetching faculty information and faculty load.
fac_list = getFacultyInfo()

# def getFacList():
#     return fac_list

load_list_copy = getFacLoad()
av_fac = set()


# setting faculty information in faculty objects
def set_load():
    global av_fac
    av_fac = set()
    load_list = getFacLoad()
    for fac in fac_list:
        fac.resetFLoad()

    while load_list:
        for l in load_list:
            for fac in fac_list:
                if fac.getFId() == l[0]:
                    l.remove(l[0])
                    if 'CS4' in l[0]:
                        sec = 'CSE 4' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                        av_fac.add(fac)

                        break
                    elif 'CS6' in l[0]:
                        sec = 'CSE 6' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                    elif 'AL4' in l[0]:
                        sec = 'AIML 4' + l[1]
                        fac.setFLoad(sec, l)
                        # print("Added")
                        load_list.remove(l)
                        av_fac.add(fac)
                        break
                    elif 'AL6' in l[0]:
                        sec = 'AIML 6' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                    elif 'CY4' in l[0]:
                        sec = 'CY 4' + l[1]
                        fac.setFLoad(sec, l)
                        # print("Added")
                        load_list.remove(l)
                        av_fac.add(fac)
                        break
                    elif 'CY6' in l[0]:
                        sec = 'CY 6' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)


# set_load()


def reset_load():
    global load_list
    load_list = getFacLoad()


# initializing reference index for Excel sheets
fac_info_row = ['16', '17', '18', '19', '20', '21', '22']

av_fac = list(av_fac)
av_fac_copy = av_fac


def reset_av_faculty():
    global av_fac
    av_fac = deepcopy(av_fac_copy)


# setting information section in the Excel sheets.
for sheet in section_sheets:
    for fac in av_fac_copy:
        if sheet in fac.getFLoad().keys():
            f_load = fac.getFLoad()[sheet]
            for i in range(len(f_load)):
                section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 2]] = f_load[i][0]
                if 'CS' in f_load[i][0]:
                    section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = cse_sub_code_names[
                        f_load[i][0]]
                elif 'AL' in f_load[i][0]:
                    section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = aiml_sub_code_names[
                        f_load[i][0]]
                elif 'CY' in f_load[i][0]:
                    section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = cy_sub_code_names[
                        f_load[i][0]]
                section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 2]] = fac.getFName()
                section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 2]] = fac.getFPhone()

# sorting faculty according to the sections
section_faculty = {}
sections = deepcopy(section_list)
sections.remove('Lib')


def get_sections():
    global sections
    return sections


for sec in sections:
    section_faculty[sec] = []

for fac in av_fac_copy:
    for sec in sections:
        if sec in fac.getFLoad():
            section_faculty[sec].append(fac)

section_faculty_copy = deepcopy(section_faculty)


def reset_sec_fac():
    global section_faculty
    section_faculty = deepcopy(section_faculty_copy)


# allotting labs to the sections
section_labs = {}
for sec in sections:
    section_labs[sec] = {}

for sec in sections:
    if 'CSE' in sec:
        for sub in cse_subjects.keys():
            if cse_subjects[sub][0] != 'l' and '02' in sub:
                l = [half_labs_copy[0], half_labs_copy[1]]
                section_labs[sec][sub + '/' + sub[0:4] + str(int(sub[4]) + 1)] = l
                # lab_load[l[0]] = lab_load[l[0]] - 2
                # lab_load[l[1]] = lab_load[l[1]] - 2
                half_labs_copy.remove(l[0])
                half_labs_copy.append(l[0])
                half_labs_copy.remove(l[1])
                half_labs_copy.append(l[1])
            elif cse_subjects[sub][0] != 'l' and '04' in sub:
                l = [half_labs_copy[0], half_labs_copy[1]]
                section_labs[sec][sub + '/' + sub[0:4] + str(int(sub[4]) + 1)] = l
                # lab_load[l[0]] = lab_load[l[0]] - 2
                # lab_load[l[1]] = lab_load[l[1]] - 2
                half_labs_copy.remove(l[0])
                half_labs_copy.append(l[0])
                half_labs_copy.remove(l[1])
                half_labs_copy.append(l[1])
            elif '06' in sub:
                section_labs[sec][sub] = ['F-11 OLD']
                # lab_load['F-11 OLD'] = lab_load['F-11 OLD'] - 1
            elif '07' in sub:
                section_labs[sec][sub] = ['F-09 OLD']
                # lab_load['F-09 OLD'] = lab_load['F-09 OLD'] - 1
    if 'AIML' in sec:
        for sub in aiml_subjects.keys():
            if aiml_subjects[sub][0] != 'l' and '02' in sub:
                l = [half_labs_copy[0], half_labs_copy[1]]
                section_labs[sec][sub + '/' + sub[0:4] + str(int(sub[4]) + 1)] = l
                # lab_load[l[0]] = lab_load[l[0]] - 2
                # lab_load[l[1]] = lab_load[l[1]] - 2
                half_labs_copy.remove(l[0])
                half_labs_copy.append(l[0])
                half_labs_copy.remove(l[1])
                half_labs_copy.append(l[1])
            elif aiml_subjects[sub][0] != 'l' and '04' in sub:
                l = [half_labs_copy[0], half_labs_copy[1]]
                section_labs[sec][sub + '/' + sub[0:4] + str(int(sub[4]) + 1)] = l
                # lab_load[l[0]] = lab_load[l[0]] - 2
                # lab_load[l[1]] = lab_load[l[1]] - 2
                half_labs_copy.remove(l[0])
                half_labs_copy.append(l[0])
                half_labs_copy.remove(l[1])
                half_labs_copy.append(l[1])
            elif '06' in sub:
                section_labs[sec][sub] = ['F-11 OLD']
                # lab_load['F-11 OLD'] = lab_load['F-11 OLD'] - 1
            elif '07' in sub:
                section_labs[sec][sub] = ['F-09 OLD']
                # lab_load['F-09 OLD'] = lab_load['F-09 OLD'] - 1
    if 'CY' in sec:
        for sub in cy_subjects.keys():
            if cy_subjects[sub][0] != 'l' and '02' in sub:
                l = [half_labs_copy[0], half_labs_copy[1]]
                section_labs[sec][sub + '/' + sub[0:4] + str(int(sub[4]) + 1)] = l
                # lab_load[l[0]] = lab_load[l[0]] - 2
                # lab_load[l[1]] = lab_load[l[1]] - 2
                half_labs_copy.remove(l[0])
                half_labs_copy.append(l[0])
                half_labs_copy.remove(l[1])
                half_labs_copy.append(l[1])
            elif cy_subjects[sub][0] != 'l' and '04' in sub:
                l = [half_labs_copy[0], half_labs_copy[1]]
                section_labs[sec][sub + '/' + sub[0:4] + str(int(sub[4]) + 1)] = l
                # lab_load[l[0]] = lab_load[l[0]] - 2
                # lab_load[l[1]] = lab_load[l[1]] - 2
                half_labs_copy.remove(l[0])
                half_labs_copy.append(l[0])
                half_labs_copy.remove(l[1])
                half_labs_copy.append(l[1])
            elif '06' in sub:
                section_labs[sec][sub] = ['F-11 OLD']
                # lab_load['F-11 OLD'] = lab_load['F-11 OLD'] - 1
            elif '07' in sub:
                section_labs[sec][sub] = ['F-09 OLD']
                # lab_load['F-09 OLD'] = lab_load['F-09 OLD'] - 1

# setting faculty for lab section/sub wise.
section_lab_faculty = {}
for sec in sections:
    section_lab_faculty[sec] = {}

for sec in sections:
    for sub in section_labs[sec].keys():
        if '06' or '07' not in sub:
            sub1 = sub[0:5]
            sub2 = sub[6:]
            l = []
            for f in section_faculty[sec]:
                for j in f.getFLoad()[sec]:
                    if j[0] == sub1:
                        l.append(f)
                        break
            for f in section_faculty[sec]:
                for j in f.getFLoad()[sec]:
                    if j[0] == sub2:
                        l.append(f)
                        break
            section_lab_faculty[sec][sub] = l
        else:
            for f in section_faculty[sec]:
                for j in f.getFLoad()[sec]:
                    if j[0] == sub:
                        section_lab_faculty[sec][sub] = [f]

# sorting faculty on the basis of section/theory
section_theory_faculty = {}
for sec in sections:
    section_theory_faculty[sec] = {}

for sec in sections:
    for fac in section_faculty[sec]:
        for sub in fac.getFLoad()[sec]:
            if int(sub[0][4]) < 6:
                section_theory_faculty[sec][sub[0]] = fac

section_theory_faculty_copy = deepcopy(section_theory_faculty)


def reset_sec_th_fac():
    global section_theory_faculty
    section_theory_faculty = deepcopy(section_theory_faculty_copy)


# mentor allotment
# mentor_list=get_sections_and_mentors(2)

# Function to check the availability of the slots.
def available_lab(sec, lab, section_tt, lab_tt, day, time):
    global section_lab_faculty, section_labs, faculty_tt

    labs_fac = section_lab_faculty[sec]
    labs = section_labs[sec]

    if section_tt[sec][day][time] == '' and section_tt[sec][day][time + 1] == '':
        pass
    else:
        # slots_copy.remove(slot)
        return False

    if sub in section_tt[sec][day]:
        return False

    if ('06' not in lab) and ('07' not in lab):
        if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
            pass
        else:
            return False
        if lab_tt[labs[lab][1]][day][time] == '' and lab_tt[labs[lab][1]][day][time + 1] == '':
            pass
        else:
            return False
        if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
            pass
        else:
            return False
        if faculty_tt[labs_fac[lab][1].getFName()][day][time] == '' and \
                faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] == '':
            pass
        else:
            return False
        if time == 0:
            fac1 = faculty_tt[labs_fac[lab][0].getFName()][day]
            fac2 = faculty_tt[labs_fac[lab][1].getFName()][day]
            if (fac1[time + 2] == '' and fac2[time + 2] == '') or (fac1[time + 3] == '' and fac2[time + 3] == '') or \
                    (fac1[time + 2] == '' and fac2[time + 3] == '') or (fac1[time + 3] == '' and fac2[time + 2] == ''):
                pass
            else:
                return False
        elif time == 2:
            fac1 = faculty_tt[labs_fac[lab][0].getFName()][day]
            fac2 = faculty_tt[labs_fac[lab][1].getFName()][day]
            if (fac1[time - 2] == '' and fac2[time - 2] == '') or (fac1[time - 1] == '' and fac2[time - 1] == '') or \
                    (fac1[time - 2] == '' and fac2[time - 1] == '') or (fac1[time - 1] == '' and fac2[time - 2] == ''):
                pass
            else:
                return False

        # cse_labs_fac[lab][0].getFLoad()[sec]

    else:
        if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
            pass
        else:
            return False
        if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
            pass
        else:
            return False
        if time == 2:
            if faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] == '' or \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time - 1] == '':
                pass
            else:
                return False
    if len(labs_fac[lab]) > 1:
        for i in labs_fac[lab]:
            if len(i.getFLoad()[sec]) > 1:
                for j in i.getFLoad()[sec]:
                    if 4 >= j[3] > 0:
                        # j[3]=j[3]-2
                        break
                    else:
                        return False
            else:
                if 4 >= i.getFLoad()[sec][0][3] > 0:
                    # i.getFLoad()[sec][0][3]=i.getFLoad()[sec][0][3]-2
                    break
                else:
                    return False
    else:
        if len(labs_fac[lab][0].getFLoad()[sec]) > 1:
            for i in labs_fac[lab][0].getFLoad()[sec]:
                if 2 >= i[3] > 0:
                    # i[3]=i[3]-2
                    break
                else:
                    return False
        else:
            if 4 >= labs_fac[lab][0].getFLoad()[sec][0][3] > 0:
                # cse_labs_fac[lab][0].getFLoad()[sec][0][3]=cse_labs_fac[lab][0].getFLoad()[sec][0][3]-2
                pass
            else:
                return False

    return True


def available_theory(sec, day, time, sub):
    global section_tt, faculty_tt, section_theory_faculty

    th_fac = section_theory_faculty[sec]

    # sec1 = tt[day][0:4]
    # sec2 = tt[day][5:]

    if section_tt[sec][day][time] == '':
        pass
    else:
        return False

    # if sub in section_tt[sec][day]:
    #     return False
    if section_tt[sec][day].count(sub) < 2:
        pass
    else:
        return False

    fac = th_fac[sub]

    if faculty_tt[fac.getFName()][day].count('') > 1:
        pass
    else:
        return False

    if faculty_tt[fac.getFName()][day][time] == '':
        fac_day = faculty_tt[fac.getFName()][day]
        if time == 0:
            if fac_day[time + 1] == '':
                pass
            elif fac_day[time + 1] == sub:
                pass
            else:
                return False

        elif time == 1:
            if fac_day[time - 1] == '' and fac_day[time + 1] == '':
                pass
            elif fac_day[time - 1] == sub and fac_day[time + 1] == '':
                pass

            elif fac_day[time - 1] == '' and fac_day[time + 1] == sub:
                pass
            elif fac_day[time + 1].__contains__('/'):
                pass
            else:
                return False

        elif time == 2:
            if fac_day[time - 1] == '' and fac_day[time + 1] == '':
                pass
            elif fac_day[time - 1] == '' and fac_day[time - 2] != sub:
                pass
            elif fac_day[time - 1] == sub and fac_day[time + 1] == '':
                pass
            elif fac_day[time - 1] == '' and fac_day[time + 1] == sub:
                pass
            elif fac_day[time - 1].__contains__('/'):
                pass
            else:
                return False

        elif time == 3:
            if fac_day[time - 1] == '':
                pass
            elif fac_day[time - 2].__contains__('/'):
                pass
            elif fac_day[time - 2] != sub:
                pass
            elif fac_day[time - 1] == sub:
                pass
            else:
                return False

        elif time == 5:
            if fac_day[time + 1] == '':
                pass
            elif fac_day[time + 1].__contains__('/'):
                pass
            elif fac_day[time + 1] == sub and fac_day[time + 2] == '':
                pass
            # elif fac_day[time + 1] == sub:
            #     pass
            elif fac_day[time + 1] == '' and fac_day[time + 2] != '':
                pass
            else:
                return False

        elif time == 6:
            if fac_day[time - 1] == '' and fac_day[time + 1] == '':
                pass
            if fac_day[time + 1] == '' and fac_day[time - 1] == sub:
                pass
            if fac_day[time + 1] == sub and fac_day[time - 1] == '':
                pass
            else:
                return True

        else:
            if fac_day[time - 1] == '' and fac_day[time - 2] == '':
                pass
            elif fac_day[time - 1] == sub and fac_day[time - 2] == '':
                pass
            # elif fac_day[time - 1] == sub:
            #     pass
            elif fac_day[time - 1] == '' and fac_day[time - 2] != '':
                pass
            else:
                return False
        pass
    # if faculty_tt[fac.getFName()][day][time] == '':
    #     pass
    else:
        return False

    load = fac.getFLoad()[sec]
    for i in load:
        if i[0] == sub:
            if 6 >= i[2] >= 1:
                break
            else:
                return False

    return True


mentor_list=get_sections_and_mentors(2)


def available_lib(sec, day, time, sub):
    global section_tt, faculty_tt, mentor_list

    fac = mentor_list[sec]

    # sec1 = tt[day][0:4]
    # sec2 = tt[day][5:]

    if section_tt[sec][day][time] == '':
        pass
    else:
        return False

    if faculty_tt[fac.getFName()][day].count('') > 1:
        pass
    else:
        return False

    if faculty_tt[fac.getFName()][day][time] == '':
        fac_day = faculty_tt[fac.getFName()][day]
        if time == 0:
            if fac_day[time + 1] == '':
                pass
            elif fac_day[time + 1] == sub:
                pass
            else:
                return False

        elif time == 1:
            if fac_day[time - 1] == '' and fac_day[time + 1] == '':
                pass
            elif fac_day[time - 1] == sub and fac_day[time + 1] == '':
                pass

            elif fac_day[time - 1] == '' and fac_day[time + 1] == sub:
                pass
            elif fac_day[time + 1].__contains__('/'):
                pass
            else:
                return False

        elif time == 2:
            if fac_day[time - 1] == '' and fac_day[time + 1] == '':
                pass
            elif fac_day[time - 1] == '' and fac_day[time - 2] != sub:
                pass
            elif fac_day[time - 1] == sub and fac_day[time + 1] == '':
                pass
            elif fac_day[time - 1] == '' and fac_day[time + 1] == sub:
                pass
            elif fac_day[time - 1].__contains__('/'):
                pass
            else:
                return False

        elif time == 3:
            if fac_day[time - 1] == '':
                pass
            elif fac_day[time - 2].__contains__('/'):
                pass
            elif fac_day[time - 2] != sub:
                pass
            elif fac_day[time - 1] == sub:
                pass
            else:
                return False

        elif time == 5:
            if fac_day[time + 1] == '':
                pass
            elif fac_day[time + 1].__contains__('/'):
                pass
            elif fac_day[time + 1] == sub and fac_day[time + 2] == '':
                pass
            # elif fac_day[time + 1] == sub:
            #     pass
            elif fac_day[time + 1] == '' and fac_day[time + 2] != '':
                pass
            else:
                return False

        elif time == 6:
            if fac_day[time - 1] == '' and fac_day[time + 1] == '':
                pass
            if fac_day[time + 1] == '' and fac_day[time - 1] == sub:
                pass
            if fac_day[time + 1] == sub and fac_day[time - 1] == '':
                pass
            else:
                return True

        else:
            if fac_day[time - 1] == '' and fac_day[time - 2] == '':
                pass
            elif fac_day[time - 1] == sub and fac_day[time - 2] == '':
                pass
            # elif fac_day[time - 1] == sub:
            #     pass
            elif fac_day[time - 1] == '' and fac_day[time - 2] != '':
                pass
            else:
                return False
        pass
    # if faculty_tt[fac.getFName()][day][time] == '':
    #     pass
    else:
        return False

    return True


# checks if any slot in any tt is empty or not
def checkEmpty(sec):
    global section_tt
    for i in section_tt[sec]:
        if '' in i:
            return True
    return False


print("started")

lab_slots = []


# setting labs
def set_labs():
    days = [0, 1, 2, 3, 4, 5]
    for sec in sections:

        count = 0
        labs_fac = section_lab_faculty[sec]
        labs = section_labs[sec]
        lab_list_copy = list(labs.keys())
        days_copy = deepcopy(days)
        unsettled = {}
        while True:
            while count < 12:
                lab = random.choice(lab_list_copy)
                day = random.choice(days_copy)
                if len(lab) > 5:
                    time = random.choice([0, 2, 6])

                else:
                    time = random.choice([2, 6])

                if lab in section_tt[sec][day]:
                    continue

                if available_lab(sec, lab, section_tt, lab_tt, day, time):
                    # print(lab)
                    # print("Passes")
                    unsettled[lab] = 0
                else:
                    if lab in unsettled.keys():
                        unsettled[lab] += 1
                    else:
                        unsettled[lab] = 1

                    if unsettled[lab] > 18:
                        return False
                    break
                if int(lab[-1]) < 6:
                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab
                    lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                    lab_tt[labs[lab][1]][day][time] = lab_tt[labs[lab][1]][day][time + 1] = sec
                    faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub
                    faculty_tt[labs_fac[lab][1].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] = sub
                    for i in labs_fac[lab]:
                        for j in range(len(i.getFLoad()[sec])):
                            if i.getFLoad()[sec][j][0] == lab[0:5] or i.getFLoad()[sec][j][0] == lab[6:]:
                                i.getFLoad()[sec][j][3] = i.getFLoad()[sec][j][3] - 2
                                break
                    c = 0
                    for i in section_tt[sec]:
                        if lab in i:
                            c = c + 1
                    if c > 1:
                        lab_list_copy.remove(lab)
                    # print(day)
                    # print(days_copy)
                    # lab_slots.append([day,time,sec,labs[lab][0],labs[lab][1],labs_fac[lab][0],labs_fac[lab][1]])
                    days_copy.remove(day)
                    count = count + 2
                    del unsettled[lab]

                else:
                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab
                    lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                    faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub
                    days_copy.remove(day)
                    lab_list_copy.remove(lab)
                    # lab_slots.append([day,time,sec,labs[lab][0],labs_fac[lab][0]])

                    count = count + 2
                    del unsettled[lab]
            if count == 12:
                break
    return True


lib_slots = []


def set_lib():
    global lib_slots
    mentor_list=get_sections_and_mentors(2)
    for sec in sections:
        c = 0
        while c < 1:
            day = random.choice(days)
            time = random.randrange(0, 8)
            fac=mentor_list[sec]
            if available_lib(sec, day, time, sub):
                lib_slots.append([day, time, sec])
                section_tt[sec][day][time] = 'M/L'
                section_tt['Lib'][day][time] = sec
                faculty_tt[fac.getFName()][day][time] = sec + '\n' + 'M/L'
                c = c + 1
                break
            else:
                continue

    return True


# set_lib()


days = [0, 1, 2, 3, 4, 5]
times = [0, 1, 2, 3, 5, 6, 7]

slot_list = []

# def last_index(day, time, fac):
#     global slot_list
#     slot_list.append([day, time, fac])


section_copy = sections.copy()


def set_theory():
    global slot_list, last_sec, lib_slots
    section_copy = sections.copy()
    for sec in section_copy:
        print(sec)
        slot_list = []
        th_fac = section_theory_faculty[sec]
        unalloted = {}
        c = 0
        while checkEmpty(sec):
            sub_list = list(th_fac.keys())
            c = c + 1
            for day in days:
                for time in [0, 1, 2, 3, 5, 6, 7]:
                    sub = random.choice(sub_list)
                    if sub not in unalloted.keys():
                        unalloted[sub] = 0
                    if available_theory(sec, day, time, sub):

                        fac = th_fac[sub]
                        slot_list.append([day, time, fac])

                        section_tt[sec][day][time] = sub
                        faculty_tt[fac.getFName()][day][time] = sub
                        for i in range(len(fac.getFLoad()[sec])):
                            if fac.getFLoad()[sec][i][0] == sub:
                                fac.getFLoad()[sec][i][2] = fac.getFLoad()[sec][i][2] - 1
                        # print("pass")
                        del unalloted[sub]

                    else:
                        unalloted[sub] = unalloted[sub] + 1

                        if unalloted[sub] > 30:
                            c = 0
                            while len(slot_list) != 0:
                                i = slot_list[0]
                                section_tt[sec][i[0]][i[1]] = ''
                                faculty_tt[i[2].getFName()][i[0]][i[1]] = ''
                                slot_list.remove(i)
                                c = c + 1

                            # c = 0
                            while len(lib_slots) != 0:
                                i = lib_slots[0]
                                section_tt[i[2]][i[0]][i[1]] = ''
                                section_tt['Lib'][i[0]][i[1]] = ''
                                lib_slots.remove(i)

                            print("removed", c)
                            return False

    return True


def setTimeTables():
    # reset_tt()
    while True:
        c2 = 0
        c = 0
        print("Setting labs")
        while True:
            if c > 0:
                reset_tt()
                copy_existing_data()
                reset_load()
                set_load()

            c = c + 1

            if set_labs():
                print("labs set")
                break
            else:
                continue
            # if set_lib():
            #     print("lib set")
            #     break
            # else:
            #     continue

        c2 = 0
        while True:
            val = 0
            c2 = c2 + 1

            if set_lib():
                print("Lib set")
                pass
            else:
                continue
            if set_theory():
                val = 1
                break

            reset_load()
            set_load()

            if c2 > 10000:
                break

        if val == 1:
            break


def copy_tt_to_excel():
    for sec in sections:
        print(sec)
        tt = section_tt[sec]
        for i in tt:
            print(i)
        sheet = section_sheets[sec]
        labs = section_labs[sec]
        labs_fac = section_lab_faculty[sec]
        th_fac = section_theory_faculty[sec]
        for day in days:
            for time in times:
                if tt[day][time] == '':
                    continue

                if tt[day][time] == 'NA' or tt[day][time] == 'Na':
                    sheet[time_cell[time] + day_cell[day]] = tt[day][time]
                    continue

                if tt[day][time] == 'M/L':
                    sheet[time_cell[time] + day_cell[day]] = 'M/L'
                    section_sheets['Lib'][time_cell[time] + day_cell[day]] = sec
                    faculty_sheets[mentor_list[sec].getFName()][day][time] = sec + '\n'+ 'M/L'
                    continue

                if tt[day][time] == sheet[time_cell[time] + day_cell[day]].value:
                    continue

                elif tt[day][time] != '' and sheet[time_cell[time] + day_cell[day]].value == None:
                    if '/' in tt[day][time] and time not in [1, 3, 7]:
                        lab = tt[day][time]
                        sheet.merge_cells(time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        sheet[time_cell[time] + day_cell[day]] = tt[day][time] + "\n" + \
                                                                 labs[tt[day][time]][0] + "/" + labs[lab][1]
                        sheet['E' + fac_info_row[int(lab[4]) - 2]] = labs[lab][0] + ' Building LNCT&S'
                        sheet['E' + fac_info_row[int(lab[10]) - 2]] = labs[lab][1] + ' Building LNCT&S'
                        lab_sheets[labs[lab][0]][time_cell[time] + day_cell[day]] = \
                            lab_sheets[labs[lab][1]][time_cell[time] + day_cell[day]] = sec
                        faculty_sheets[labs_fac[lab][0].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][1].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][0].getFName()][time_cell[time] + day_cell[day]] = \
                            faculty_sheets[labs_fac[lab][1].getFName()][
                                time_cell[time] + day_cell[day]] = sec + "\n" + lab + "\n" + labs[lab][0] + '/' + \
                                                                   labs[lab][
                                                                       1]
                    elif int(tt[day][time][4]) > 5:
                        if time in [0, 2, 6]:
                            pass
                        else:
                            continue
                        labs_fac = section_lab_faculty[sec]
                        lab = tt[day][time]
                        sheet.merge_cells(time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        sheet[time_cell[time] + day_cell[day]] = tt[day][time] + "\n" + labs[lab][0]
                        sheet['E' + fac_info_row[int(lab[4]) - 2]] = labs[lab][0] + ' Building LNCTS'
                        lab_sheets[labs[lab][0]][time_cell[time] + day_cell[day]] = sec
                        faculty_sheets[labs_fac[lab][0].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][0].getFName()][
                            time_cell[time] + day_cell[day]] = sec + "\n" + lab + '\n' + labs[lab][0]
                    else:
                        sub = tt[day][time]
                        if '/' not in sub:
                            pass
                        else:
                            continue

                        fac = th_fac[sub]
                        sheet[time_cell[time] + day_cell[day]] = tt[day][time]
                        faculty_sheets[fac.getFName()][time_cell[time] + day_cell[day]] = sec + "\n" + tt[day][time]


# Do not touch the code below.
# Counting the number of folders/files present in the directory and then creating a new directory on the basis of that
def set_sheets():
    files = next(os.walk('C:/Users/Swadesh Sharma/Desktop/SchedulXpert/'))[1]
    if not files:
        n = 0
    else:
        n = len(files)
    newpath = 'C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test ' + str(n + 1)

    # Creating new directories
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        os.makedirs(newpath + '/Faculty')
        os.makedirs(newpath + '/Sections')
        os.makedirs(newpath + '/Labs')

    # saving the files in the directory creates
    for i in workbooks.keys():
        if 'Lib' in i or 'CS' in i or 'AIML' in i or 'CY' in i:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Sections/" + i + ".xlsx")
        elif 'Dr' in i or 'Prof' in i:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Faculty/" + i + ".xlsx")
        else:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Labs/" + i + ".xlsx")


#
# setTimeTables()
# copy_tt_to_excel()
# set_sheets()
print("finished")
