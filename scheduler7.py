from copy import deepcopy
import random
import os
from openpyxl.cell.cell import MergedCell
from random import choice

import openpyxl
import get_info


def getWorkBook():
    return openpyxl.load_workbook("sheets/Book.xlsx")


def getWorkBookLab():
    return openpyxl.load_workbook("sheets/BookLab.xlsx")


def getTT():
    TT = [['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', ''],
          ['', '', '', '', 'Lunch', '', '', ''], ['', '', '', '', 'Lunch', '', '', ''],
          ['', '', '', '', 'Lunch', '', '', '']]
    return TT


# days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
day_cell = ['7', '8', '9', '10', '11']
time_cell = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

# sections
sections = get_info.get_sections_and_mentors(1)
sec_list = sections.keys()

# initializing workbooks
workbooks = {
    'Lib': getWorkBook()
}

for sec in sec_list:
    workbooks[sec] = getWorkBook()

# initializing sections and mentors
mentors = sections

# getting faculty information
faculty = get_info.getFacultyInfo()
# fac_list = []

for fac in faculty:
    workbooks[fac.getFName()] = getWorkBook()
    # fac_list.append(fac.getFName())

# getting labs
labs = get_info.get_labs()
# print(labs)
for lab in labs.keys():
    workbooks[lab] = getWorkBookLab()

# Dictionary containing the name of the sec, labs and faculty with their respective worksheets
section_sheets = {}
lab_sheets = {}
faculty_sheets = {}

for sec in sections.keys():
    section_sheets[sec] = workbooks[sec].active

section_sheets['Lib'] = workbooks['Lib'].active

for lab in labs.keys():
    lab_sheets[lab] = workbooks[lab].active

for fac in faculty:
    faculty_sheets[fac.getFName()] = workbooks[fac.getFName()].active


# setting load
def set_load():
    global av_fac
    av_fac = set()
    load_list = get_info.getFacLoad()
    for fac in faculty:
        fac.resetFLoad()

    while load_list:
        for l in load_list:
            for fac in faculty:
                if fac.getFId() == l[0]:
                    l.remove(l[0])
                    if 'CS3' in l[0]:
                        sec = 'CSE 3' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                        av_fac.add(fac)

                        break
                    elif 'CS5' in l[0]:
                        sec = 'CSE 5' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                    elif 'AL3' in l[0]:
                        sec = 'AIML 3' + l[1]
                        fac.setFLoad(sec, l)
                        # print("Added")
                        load_list.remove(l)
                        av_fac.add(fac)
                        break
                    elif 'AL5' in l[0]:
                        sec = 'AIML 5' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)
                    elif 'CY3' in l[0]:
                        sec = 'CY 3' + l[1]
                        fac.setFLoad(sec, l)
                        # print("Added")
                        load_list.remove(l)
                        av_fac.add(fac)
                        break
                    elif 'CY5' in l[0]:
                        sec = 'CY 5' + l[1]
                        fac.setFLoad(sec, l)
                        load_list.remove(l)


set_load()


def reset_load():
    global load_list
    load_list = get_info.getFacLoad()


# Dictionary containing the name of the sec, lab and faculty with their timetable structure
section_tt = {}
lab_tt = {}
faculty_tt = {}

for sec in sections.keys():
    section_tt[sec] = getTT()

section_tt['Lib'] = getTT()

for lab in labs.keys():
    lab_tt[lab] = getTT()

for fac in faculty:
    faculty_tt[fac.getFName()] = getTT()

all_tt = {
    'Faculty': faculty_tt,
    'Labs': lab_tt,
    'Sections': section_tt
}

all_presets = {
    'Faculty': [],
    'Labs': [],
    'Sections': []
}


def reset_presets():
    for i in all_presets.keys():
        all_presets[i] = []


section_labs = {}
section_labs_fac = {}
section_th = {}
section_th_fac = {}
for sec in sections.keys():
    section_labs[sec] = {}
    section_labs_fac[sec] = {}

for sec in sections.keys():
    if '3' in sec:
        subjects = get_info.get_sub_by_section(sec)
        l = []
        section_th[sec] = []
        section_th_fac[sec] = {}
        for sub in subjects.keys():
            if subjects[sub][1] == 'B':
                l.append(sub)
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []


            elif subjects[sub][1] == 'L':
                l.append(sub)

            elif subjects[sub][1] == 'T':
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

        section_labs[sec][l[0] + '/' + l[1]] = []
        section_labs[sec][l[2] + '/' + l[3]] = []
        section_labs[sec][l[4]] = []

        section_labs_fac[sec][l[0] + '/' + l[1]] = []
        section_labs_fac[sec][l[2] + '/' + l[3]] = []
        section_labs_fac[sec][l[4]] = []
    elif '5' in sec:
        subjects = get_info.get_sub_by_section(sec)
        # print(subjects)
        section_th[sec] = []
        section_th_fac[sec] = {}
        for sub in subjects.keys():
            if subjects[sub][1] == 'B':
                section_labs[sec][sub] = []
                section_labs_fac[sec][sub] = []
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

            elif subjects[sub][1] == 'L':
                section_labs[sec][sub] = []
                section_labs_fac[sec][sub] = []

            elif subjects[sub][1] == 'T':
                section_th[sec].append(sub)
                section_th_fac[sec][sub] = []

# for i, j in section_th.items():
#     print(i, j)

for sec in sections.keys():
    for sub in section_labs[sec].keys():
        if '/' in sub:
            s1 = sub[0:5]
            s2 = sub[6:]
            l = ['', '']
            for fac in faculty:
                if sec in fac.getFLoad().keys():
                    load = fac.getFLoad()[sec]
                else:
                    continue
                for i in load:
                    if i[0] == s1:
                        l[0] = fac
                        break
                for i in load:
                    if i[0] == s2:
                        l[1] = fac
                        break
            section_labs_fac[sec][sub] = l
        else:
            for fac in faculty:
                if sec in fac.getFLoad().keys():
                    load = fac.getFLoad()[sec]
                else:
                    continue
                comp = 0
                for i in load:

                    if i[0] == sub:
                        section_labs_fac[sec][sub].append(fac)
                        comp = 1
                        break
                if comp == 1:
                    break

for sec in sec_list:
    for sub in section_th[sec]:
        for fac in faculty:
            if sec in fac.getFLoad().keys():
                load = fac.getFLoad()[sec]
            else:
                continue
            comp = 0
            for i in load:
                if i[0] == sub:
                    section_th_fac[sec][sub] = fac
                    comp = 1
            if comp == 1:
                break

# print(section_labs_fac)
# for sec in sections.keys():
#     print(sec)
#     for i,j in section_labs_fac[sec].items():
#         if '/' in i:
#             print(i,j[0].getFName(),j[1].getFName())
#         else:
#             print(i, j[0].getFName())


# setting regular information in the sheets.
for sheet in section_sheets.keys():
    if sheet == 'Lib':
        section_sheets['Lib']['A4'] = 'Branch: CS/AI/CY'
        section_sheets['Lib']['C4'] = 'Sections: ALL'
        section_sheets['Lib']['F7'] = 'LUNCH'
        continue
    section_sheets[sheet]['A4'] = 'Branch: ' + sheet[0:-2]
    section_sheets[sheet]['C4'] = 'Section: ' + sheet[-2:]
    section_sheets[sheet]['F7'] = 'LUNCH'

for sheet in lab_sheets.keys():
    lab_sheets[sheet]['A4'] = 'Branch: ALL'
    lab_sheets[sheet]['C4'] = 'Section: ALL'
    lab_sheets[sheet]['F7'] = 'LUNCH'
    lab_sheets[sheet]['E4'] = 'Room No.: ' + sheet + " S"
    lab_sheets[sheet].title = sheet

for fac in faculty:
    faculty_sheets[fac.getFName()]['A4'] = 'Name: ' + fac.getFName()
    faculty_sheets[fac.getFName()]['F7'] = 'LUNCH'
    faculty_sheets[fac.getFName()]['E4'] = ''

# initializing reference index for Excel sheets
fac_info_row = ['14', '15', '16', '17', '18', '19', '20', '21']

# setting information section in the Excel sheets.
for sheet in section_sheets:
    subjects = get_info.get_sub_by_section(sheet)
    if 'Lib' in sheet:
        continue
    section_sheets[sheet]['G27'] = mentors[sheet].getFName()
    section_sheets[sheet]['I27'] = mentors[sheet].getFPhone()
    # print(subjects)
    for fac in faculty:
        if sheet in fac.getFLoad().keys():
            f_load = fac.getFLoad()[sheet]
            if '3' in sheet:
                for i in range(len(f_load)):
                    # print(fac_info_row[int(f_load[i][0][4]) - 2])
                    section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 1]] = f_load[i][0]
                    if 'CS' in f_load[i][0]:
                        section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = subjects[f_load[i][0]][0]
                    elif 'AL' in f_load[i][0]:
                        section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = subjects[f_load[i][0]][0]
                    elif 'CY' in f_load[i][0]:
                        section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = subjects[f_load[i][0]][0]
                    section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFName()
                    section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFPhone()

            elif '5' in sheet:
                for i in range(len(f_load)):
                    if '8' in f_load[i][0]:
                        section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 2]] = f_load[i][0]
                        if 'CS' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = \
                                subjects[f_load[i][0]][0]
                        elif 'AL' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = \
                                subjects[f_load[i][0]][0]
                        elif 'CY' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 2]] = \
                                subjects[f_load[i][0]][0]
                        section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 2]] = fac.getFName()
                        section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 2]] = fac.getFPhone()
                    else:
                        section_sheets[sheet]['A' + fac_info_row[int(f_load[i][0][4]) - 1]] = f_load[i][0]
                        if 'CS' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = \
                                subjects[f_load[i][0]][0]
                        elif 'AL' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = \
                                subjects[f_load[i][0]][0]
                        elif 'CY' in f_load[i][0]:
                            section_sheets[sheet]['B' + fac_info_row[int(f_load[i][0][4]) - 1]] = \
                                subjects[f_load[i][0]][0]
                        section_sheets[sheet]['G' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFName()
                        section_sheets[sheet]['I' + fac_info_row[int(f_load[i][0][4]) - 1]] = fac.getFPhone()

# lab allotment
lab_allotment = get_info.get_lab_allotment()
# for i,j in lab_allotment.items():
#     print(i,j)

for sec in sections.keys():
    if '3' in sec:
        for sub in section_labs[sec].keys():
            if '/' in sub:
                s1 = sub[0:5]
                s2 = sub[6:]
                li = ['', '']
                l = lab_allotment[s1]
                for i in l:
                    if sec[-1] == i[0]:
                        li[0] = i[1]
                        break
                l = lab_allotment[s2]
                for i in l:
                    if sec[-1] == i[0]:
                        li[1] = i[1]
                        break
                section_labs[sec][sub] = li
            else:
                l = lab_allotment[sub]
                for i in l:
                    if sec[-1] == i[0]:
                        section_labs[sec][sub] = [i[1]]
                        break
    elif '5' in sec:
        for sub in section_labs[sec].keys():
            l = lab_allotment[sub]
            for i in l:
                if sec[-1] == i[0]:
                    section_labs[sec][sub] = [i[1]]
                    break

section_tt_copy = deepcopy(section_tt)
lab_tt_copy = deepcopy(lab_tt)
faculty_tt_copy = deepcopy(faculty_tt)

all_tt = {
    'Faculty': faculty_tt,
    'Labs': lab_tt,
    'Sections': section_tt
}

all_presets = {
    'Faculty': [],
    'Labs': [],
    'Sections': []
}


# copy_data
# def copy_prev_data():
#     # reset_tt()
#     for sec in section_tt.keys():
#         tt = section_tt_copy[sec]
#         for day in range(len(tt)):
#             for time in range(len(tt[day])):
#                 tt[day][time] = section_tt[sec][day][time]
#
#     for fac in faculty_tt.keys():
#         tt = faculty_tt_copy[fac]
#         for day in range(len(tt)):
#             for time in range(len(tt[day])):
#                 tt[day][time] = faculty_tt[fac][day][time]
#
#     for lab in lab_tt.keys():
#         tt = lab_tt_copy[lab]
#         for day in range(len(tt)):
#             for time in range(len(tt[day])):
#                 tt[day][time] = lab_tt[lab][day][time]


def reset_tt():
    global all_presets, section_tt, lab_tt, faculty_tt, section_tt_copy, lab_tt_copy, faculty_tt_copy
    section_tt = deepcopy(section_tt_copy)
    lab_tt = deepcopy(lab_tt_copy)
    faculty_tt = deepcopy(faculty_tt_copy)
    presets = all_presets['Sections']
    for i in presets:
        sec = i[2]
        day = i[0]
        time = i[1]
        data = i[3]
        section_tt[sec][day][time] = data
    # for type in all_presets.keys():
    #
    # if type == 'Faculty':
    #     for i in presets:
    #         fac = i[2]
    #         day = i[0]
    #         time = i[1]
    #         data = i[3]
    #         faculty_tt[fac][day][time] = data
    # elif type == 'Labs':
    #     for i in presets:
    #         lab = i[2]
    #         day = i[0]
    #         time = i[1]
    #         data = i[3]
    #         lab_tt[lab][day][time] = data
    # elif type == 'Sections':
    #     for i in presets:
    #         sec = i[2]
    #         day = i[0]
    #         time = i[1]
    #         data = i[3]
    #         section_tt[sec][day][time] = data


# for sec in sections.keys():
#     print(sec)
#     for i, j in section_th_fac[sec].items():
#         print(i, j.getFName())

# CS302/CS304 0,2,6

def available_lab(sec, lab, section_tt, lab_tt, day, time):
    global section_labs_fac, section_labs, faculty_tt

    labs_fac = section_labs_fac[sec]
    labs = section_labs[sec]

    if section_tt[sec][day][time] == section_tt[sec][day][time + 1] == '':
        pass
    else:
        return False

    if sub in section_tt[sec][day]:
        return False

    if '3' in sec:
        if '07' not in lab:
            if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
                pass
            else:
                return False
            if lab_tt[labs[lab][1]][day][time] == '' and lab_tt[labs[lab][1]][day][time + 1] == '':
                pass
            else:
                return False
            if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
                pass
            else:
                return False
            if faculty_tt[labs_fac[lab][1].getFName()][day][time] == '' and \
                    faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] == '':
                pass
            else:
                return False
            if time == 0:
                fac1 = faculty_tt[labs_fac[lab][0].getFName()][day]
                fac2 = faculty_tt[labs_fac[lab][1].getFName()][day]
                if (fac1[time + 2] == '' and fac1[time + 3] != '' and fac2[time + 2] == '' and fac2[time + 3] != '') or \
                        (fac1[time + 3] == '' and fac1[time + 2] != '' and fac2[time + 3] == '' and fac2[
                            time + 2] != '') or \
                        (fac1[time + 2] == '' and fac1[time + 3] != '' and fac2[time + 3] == '' and fac2[
                            time + 2] != '') or \
                        (fac1[time + 3] == '' and fac1[time + 2] != '' and fac2[time + 2] == '' and fac2[
                            time + 3] != '') or \
                        (fac1[time + 2] == '' and fac1[time + 3] == '' and fac2[time + 2] == '' and fac2[
                            time + 3] == ''):
                    pass
                else:
                    return False
            elif time == 2:
                fac1 = faculty_tt[labs_fac[lab][0].getFName()][day]
                fac2 = faculty_tt[labs_fac[lab][1].getFName()][day]
                if (fac1[time - 2] == '' and fac1[time - 1] != '' and fac2[time - 2] == '' and fac2[time - 1] != '') or \
                        (fac1[time - 1] == '' and fac1[time - 2] != '' and fac2[time - 1] == '' and fac2[
                            time - 2] != '') or \
                        (fac1[time - 2] == '' and fac1[time - 1] != '' and fac2[time - 1] == '' and fac2[
                            time - 2] != '') or \
                        (fac1[time - 1] == '' and fac1[time - 2] != '' and fac2[time - 2] == '' and fac2[
                            time - 1] != '') or \
                        (fac1[time - 2] == '' and fac1[time - 1] == '' and fac2[time - 2] == '' and fac2[
                            time - 1] == ''):
                    pass
                else:
                    return False

            # cse_labs_fac[lab][0].getFLoad()[sec]

        else:
            if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
                pass
            else:
                return False
            if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
                pass
            else:
                return False

            if time == 0:
                if faculty_tt[labs_fac[lab][0].getFName()][day][time + 2] == '' or \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 3] == '':
                    pass
                else:
                    return False

            if time == 2:
                if faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] == '' or \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time - 1] == '':
                    pass
                else:
                    return False
    elif '5' in sec:
        if lab_tt[labs[lab][0]][day][time] == '' and lab_tt[labs[lab][0]][day][time + 1] == '':
            pass
        else:
            return False
        if faculty_tt[labs_fac[lab][0].getFName()][day][time] == '' and \
                faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] == '':
            pass
        else:
            return False

        if time == 0:
            if faculty_tt[labs_fac[lab][0].getFName()][day][time + 2] == '' or \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time + 3] == '':
                pass
            else:
                return False

        if time == 2:
            if faculty_tt[labs_fac[lab][0].getFName()][day][time - 2] == '' or \
                    faculty_tt[labs_fac[lab][0].getFName()][day][time - 1] == '':
                pass
            else:
                return False

    return True


slots = []
visited = {}


def visited_dict():
    for sec in sec_list:
        visited[sec] = False


visited_dict()


def reset_visited():
    for i in visited.keys():
        visited[i] = False


last_sec = ''

# lab_count={}
sem_3_sec = []
sem_5_sec = []
for i in sec_list:
    if '3' in i:
        sem_3_sec.append(i)
    elif '5' in i:
        sem_5_sec.append(i)
# sem_3_sec.reverse()
# sem_5_sec.reverse()
print(sem_3_sec, sem_5_sec)

lab_counter = {}


def reset_lab_counter():
    for lab in lab_tt.keys():
        lab_counter[lab] = 0


reset_lab_counter()

lab_day_counter = {}

for sec in sec_list:
    lab_day_counter[sec] = {}
    for day in [0, 1, 2, 3, 4]:
        lab_day_counter[sec][day] = 2


def reset_lab_day_counter():
    for sec in lab_day_counter.keys():
        for day in lab_day_counter[sec].keys():
            lab_day_counter[sec][day] = 2


def reset_lab_day_counter_sec(sec):
    for day in lab_day_counter[sec].keys():
        lab_day_counter[sec][day] = 2


print(lab_day_counter)


# prev_slot=[]
def set_labs():
    global section_tt, faculty_tt, lab_tt, last_sec, lab_counter
    days = [0, 1, 2, 3, 4]
    for sec in sem_5_sec:
        # if '5' in sec:
        #     pass
        # else:
        #     continue
        #
        if visited[sec]:
            continue

        # print(sec)
        slots = []
        # all_slots=[]
        # for day in [0,1,2,3,4]:
        #     for time in [0,2,6]:
        #         all_slots.append([day,time])

        count = 0
        labs_fac = section_labs_fac[sec]
        labs = section_labs[sec]
        lab_list_copy = list(labs.keys())
        lab_count = {}
        for lab in lab_list_copy:
            if '/' in lab:
                lab_count[lab] = 4
            else:
                lab_count[lab] = 2
        # print(lab_count)
        days_copy = deepcopy(days)
        unsettled = {}
        while True:
            while count < 10:
                lab = random.choice(lab_list_copy)
                # s=random.choice(all_slots)
                day = random.choice(days_copy)
                # day = random.choice(days_copy)
                if 'F-11 OLD' not in labs[lab]:
                    time = random.choice([0, 2, 6])
                else:
                    time = random.choice([2, 6])
                # if 'F-11 OLD' in labs[lab]:
                #     if s[1]==0:
                #         continue
                #     else:
                #         time=s[1]
                # else:
                #     time=s[1]

                # if lab in section_tt[sec][day]:
                #     continue

                if available_lab(sec, lab, section_tt, lab_tt, day, time):
                    # print('passed')
                    # visited_lab_slots[sec][lab].append(s)
                    # all_slots.remove(s)
                    pass
                else:
                    if lab in unsettled.keys():
                        unsettled[lab] += 1
                    else:
                        unsettled[lab] = 1

                    if unsettled[lab] > 18:
                        # print('Returning')
                        for slot in slots:
                            day = slot[0]
                            time = slot[1]
                            sec = slot[4]
                            if len(slot[2]) == 2:
                                lab_tt[slot[2][0]][day][time] = lab_tt[slot[2][0]][day][time + 1] = \
                                    faculty_tt[slot[3][0].getFName()][day][time] = \
                                    faculty_tt[slot[3][0].getFName()][day][time + 1] = \
                                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = \
                                    lab_tt[slot[2][1]][day][time] = lab_tt[slot[2][1]][day][time + 1] = \
                                    faculty_tt[slot[3][1].getFName()][day][time] = \
                                    faculty_tt[slot[3][1].getFName()][day][time + 1] = ''
                                lab_counter[slot[2][0]] = lab_counter[slot[2][0]] - 1
                                lab_counter[slot[2][1]] = lab_counter[slot[2][1]] - 1
                            else:
                                lab_tt[slot[2][0]][day][time] = lab_tt[slot[2][0]][day][time + 1] = \
                                    faculty_tt[slot[3][0].getFName()][day][time] = \
                                    faculty_tt[slot[3][0].getFName()][day][time + 1] = \
                                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = ''
                                lab_counter[slot[2][0]] = lab_counter[slot[2][0]] - 1
                        last_sec = sec
                        # for s in visited.keys():
                        #     if '5' in s and sec.__contains__(s[0:2]):
                        #         visited[s] = False
                        visited[sec] = False
                        reset_lab_day_counter_sec(sec)
                        # slots.clear()
                        # last_sec=sec
                        return False
                    break

                if '/' in lab:
                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab
                    lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                    lab_tt[labs[lab][1]][day][time] = lab_tt[labs[lab][1]][day][time + 1] = sec
                    faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = lab
                    faculty_tt[labs_fac[lab][1].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] = lab
                    slots.append([day, time, labs[lab], labs_fac[lab], sec])

                    lab_count[lab] = lab_count[lab] - 2
                    lab_counter[labs[lab][0]] = lab_counter[labs[lab][0]] + 1
                    lab_counter[labs[lab][1]] = lab_counter[labs[lab][1]] + 1
                    if lab_count[lab] == 0:
                        lab_list_copy.remove(lab)
                    all_presets['Sections'].append([day, time, sec, lab])
                    all_presets['Sections'].append([day, time + 1, sec, lab])
                    lab_day_counter[sec][day] -= 1
                    if lab_day_counter[sec][day] == 0:
                        days_copy.remove(day)
                    # days_copy.remove(day)
                    count = count + 2
                    # del unsettled[lab]
                    # settled = True
                    break

                else:
                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab + 'LAB'
                    lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                    faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub + 'LAB'
                    slots.append([day, time, labs[lab], labs_fac[lab], sec])
                    # days_copy.remove(day)
                    lab_day_counter[sec][day] -= 1
                    if lab_day_counter[sec][day] == 0:
                        days_copy.remove(day)
                    lab_count[lab] = lab_count[lab] - 2
                    lab_counter[labs[lab][0]] = lab_counter[labs[lab][0]] + 1
                    if lab_count[lab] == 0:
                        lab_list_copy.remove(lab)
                    all_presets['Sections'].append([day, time, sec, lab])
                    all_presets['Sections'].append([day, time + 1, sec, lab])
                    count = count + 2
                    # del unsettled[lab]
                    # settled = True
                    break

            # lab_list_copy = list(labs.keys())
            if count == 10:
                visited[sec] = True
                # sec_l.remove(sec)
                # for i in section_tt[sec]:
                #     print(i)
                break

    for sec in sem_3_sec:
        if 'AIML 3C' == sec:
            continue
        # else:
        #     continue
        #

        if visited[sec] == True:
            continue

        # print(sec)
        # if sec == 'CSE 3A' or sec == 'CY 3A' or sec == 'AIML 3A':
        #     slots = []
        # else:
        #     pass
        slots = []
        # all_slots = []
        # for day in [0, 1, 2, 3, 4]:
        #     for time in [0, 2, 6]:
        #         all_slots.append([day, time])
        count = 0
        labs_fac = section_labs_fac[sec]
        labs = section_labs[sec]
        lab_list_copy = list(labs.keys())
        lab_count = {}
        for lab in lab_list_copy:
            if '/' in lab:
                lab_count[lab] = 4
            else:
                lab_count[lab] = 2
        # print(lab_count)
        days_copy = deepcopy(days)
        unsettled = {}
        while True:
            while count < 10:
                lab = random.choice(lab_list_copy)
                # s=random.choice(all_slots)
                day = random.choice(days_copy)
                # day=s[0]
                # if 'F-11 OLD' in labs[lab]:
                #     if s[1]==0:
                #         continue
                #     else:
                #         time=s[1]
                # else:
                #     time=s[1]
                if 'F-11 OLD' not in labs[lab]:
                    time = random.choice([0, 2, 6])
                else:
                    time = random.choice([2, 6])

                # if lab in section_tt[sec][day]:
                #     continue

                if available_lab(sec, lab, section_tt, lab_tt, day, time):
                    # print('passed')
                    # visited_lab_slots[sec][lab].append(s)
                    # all_slots.remove(s)
                    pass
                else:
                    if lab in unsettled.keys():
                        unsettled[lab] += 1
                    else:
                        unsettled[lab] = 1

                    if unsettled[lab] > 18:
                        # print('Returning', len(slots))
                        for slot in slots:

                            day = slot[0]
                            time = slot[1]
                            sec = slot[4]
                            if len(slot[2]) == 2:
                                lab_tt[slot[2][0]][day][time] = lab_tt[slot[2][0]][day][time + 1] = \
                                    faculty_tt[slot[3][0].getFName()][day][time] = \
                                    faculty_tt[slot[3][0].getFName()][day][time + 1] = \
                                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = \
                                    lab_tt[slot[2][1]][day][time] = lab_tt[slot[2][1]][day][time + 1] = \
                                    faculty_tt[slot[3][1].getFName()][day][time] = \
                                    faculty_tt[slot[3][1].getFName()][day][time + 1] = ''
                                lab_counter[slot[2][0]] = lab_counter[slot[2][0]] - 1
                                lab_counter[slot[2][1]] = lab_counter[slot[2][1]] - 1
                            else:
                                lab_tt[slot[2][0]][day][time] = lab_tt[slot[2][0]][day][time + 1] = \
                                    faculty_tt[slot[3][0].getFName()][day][time] = \
                                    faculty_tt[slot[3][0].getFName()][day][time + 1] = \
                                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = ''
                                lab_counter[slot[2][0]] = lab_counter[slot[2][0]] - 1

                        last_sec = sec
                        # for s in visited.keys():
                        #     if '3' in s and sec.__contains__(s[0:2]):
                        #         visited[s] = False
                        visited[sec] = False
                        reset_lab_day_counter_sec(sec)
                        # slots.clear()

                        return False
                    break

                if '/' in lab:
                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab
                    lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                    lab_tt[labs[lab][1]][day][time] = lab_tt[labs[lab][1]][day][time + 1] = sec
                    faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub
                    faculty_tt[labs_fac[lab][1].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][1].getFName()][day][time + 1] = sub
                    slots.append([day, time, labs[lab], labs_fac[lab], sec])

                    lab_count[lab] = lab_count[lab] - 2
                    lab_counter[labs[lab][0]] = lab_counter[labs[lab][0]] + 1
                    lab_counter[labs[lab][1]] = lab_counter[labs[lab][1]] + 1
                    if lab_count[lab] == 0:
                        lab_list_copy.remove(lab)
                    # days_copy.remove(day)
                    count = count + 2
                    all_presets['Sections'].append([day, time, sec, lab])
                    all_presets['Sections'].append([day, time + 1, sec, lab])
                    lab_day_counter[sec][day] -= 1
                    if lab_day_counter[sec][day] == 0:
                        days_copy.remove(day)
                    # del unsettled[lab]
                    # settled = True
                    break

                else:
                    section_tt[sec][day][time] = section_tt[sec][day][time + 1] = lab + 'LAB'
                    lab_tt[labs[lab][0]][day][time] = lab_tt[labs[lab][0]][day][time + 1] = sec
                    faculty_tt[labs_fac[lab][0].getFName()][day][time] = \
                        faculty_tt[labs_fac[lab][0].getFName()][day][time + 1] = sub + 'LAB'
                    slots.append([day, time, labs[lab], labs_fac[lab], sec])
                    # days_copy.remove(day)
                    all_presets['Sections'].append([day, time, sec, lab])
                    all_presets['Sections'].append([day, time + 1, sec, lab])
                    lab_day_counter[sec][day] -= 1
                    if lab_day_counter[sec][day] == 0:
                        days_copy.remove(day)
                    # lab_count[lab] = lab_count[lab] - 2
                    # lab_counter[labs[lab][0]] = lab_counter[labs[lab][0]] + 1
                    # if lab_count[lab] == 0:
                    #     lab_list_copy.remove(lab)
                    count = count + 2

                    break

            if count == 10:
                visited[sec] = True
                break
        # comp=1
    # slots.clear()

    return True


def available_lib(day, time, sec, fac):
    if section_tt[sec][day][time] == '':
        pass
    else:
        return False

    f = fac.getFName()
    if faculty_tt[f][day][time] == '':
        pass
    else:
        return False

    if section_tt['Lib'][day][time] == '':
        pass
    else:
        return False

    return True


lib_slots = {}


def set_lib():
    days = [0, 1, 2, 3, 4]
    for sec in sem_5_sec:
        c = 0
        while c < 1:
            day = random.choice(days)
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            fac = mentors[sec]
            if available_lib(day, time, sec, fac):
                lib_slots[sec] = [day, time, sec]
                section_tt[sec][day][time] = 'M/L'
                section_tt['Lib'][day][time] = sec
                faculty_tt[fac.getFName()][day][time] = sec + '\n' + 'M/L'
                c = c + 1
                break
            # else:
            #     continue
    for sec in sem_3_sec:
        c = 0
        while c < 1:
            day = random.choice(days)
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            fac = mentors[sec]
            if available_lib(day, time, sec, fac):
                lib_slots[sec] = [day, time, sec, fac]
                section_tt[sec][day][time] = 'M/L'
                section_tt['Lib'][day][time] = sec
                faculty_tt[fac.getFName()][day][time] = sec + '\n' + 'M/L'
                c = c + 1
                break
            else:
                continue


def checkEmpty(sec):
    global section_tt
    for i in section_tt[sec]:
        # print(sec)
        if '' in i:
            # tt=section_tt[sec]
            # for i in tt:
            #     print(i)
            return True
    return False


# if faculty_tt[fac.getFName()][day].count('') > 1:
#     pass
# else:
#     return False
# if faculty_tt[fac.getFName()][day][time] == '' and '3' in sec:
#     fac_day = faculty_tt[fac.getFName()][day]
#     if time == 0:
#         if fac_day[time + 1] == fac_day[time + 2] == fac_day[time + 3] == '':
#             pass
#         elif fac_day[1] == '' and (fac_day[2].__contains__('/') or fac_day[2].__contains__('LAB')):
#             pass
#         elif (fac_day[time + 1] == fac_day[time + 3] == '') and fac_day[time + 2] != '':
#             pass
#         elif fac_day[time + 1] != '' and fac_day[time + 2] == '' and fac_day[time + 3] != '':
#             pass
#         elif fac_day[time + 1] == '' and fac_day[time + 2] != '' and fac_day[time + 3] != '':
#             pass
#         else:
#             return False
#
#     elif time == 1:
#         if fac_day[time - 1] == '' and fac_day[time + 1] == '' and fac_day[time + 2] == '':
#             pass
#         elif fac_day[time - 1] == fac_day[time + 2] == '' and fac_day[time + 1] != '':
#             pass
#         elif fac_day[0] != '' and fac_day[2] == '' and fac_day[3] == '':
#             pass
#         elif fac_day[0] != '' and fac_day[3] != '' and fac_day[2] == '':
#             pass
#         elif fac_day[0] == '' and (fac_day[2].__contains__('/') or fac_day[2].__contains__('LAB')):
#             pass
#         else:
#             return False
#
#     elif time == 2:
#         if fac_day[0] == fac_day[1] == fac_day[3] == '':
#             pass
#         elif (fac_day[0].__contains__('/') or fac_day[0].__contains__('LAB')) and fac_day[3] == '':
#             pass
#         elif fac_day[0] != '' and fac_day[1] == '' and fac_day[3] != '':
#             pass
#         elif fac_day[0] == fac_day[1] == '' and fac_day[3] != '':
#             pass
#         else:
#             return True
#
#     elif time == 3:
#         if fac_day[0] == fac_day[1] == fac_day[2] == '':
#             pass
#         elif (fac_day[0].__contains__('/') or fac_day[0].__contains__('LAB')) and fac_day[2] == '':
#             pass
#         elif fac_day[0] == fac_day[1] == '' and fac_day[2] != '':
#             pass
#         elif fac_day[0] != '' and fac_day[1] != '' and fac_day[2] == '':
#             pass
#         elif fac_day[0] == '' and fac_day[2] == '' and fac_day[1] != '':
#             pass
#         elif fac_day[0] != '' and fac_day[2] != '' and fac_day[1] == '':
#             pass
#         else:
#             return False
#
#     elif time == 5:
#         if fac_day[6] == '' and fac_day[7] == '':
#             pass
#         elif fac_day[6] == '' and fac_day[7] != '':
#             pass
#         elif fac_day[6] != '' and fac_day[7] == '':
#             pass
#         elif fac_day[6].__contains__('/') or fac_day[6].__contains__('Lab'):
#             pass
#         else:
#             return False
#
#     elif time == 6:
#         if fac_day[5] == fac_day[7] =='':
#             pass
#         elif fac_day[5]=='' and fac_day[7]!='':
#             pass
#         elif fac_day[5]!='' and fac_day[7]=='':
#             pass
#         else:
#             return False
#
#     else:
#         if fac_day[5]==fac_day[6]=='':
#             pass
#         elif fac_day[5]=='' and fac_day[6]!='':
#             pass
#         elif fac_day[5]!='' and fac_day[6]=='':
#             pass
#         else:
#             return False
#     pass
# # if faculty_tt[fac.getFName()][day][time] == '':
# #     pass
# elif faculty_tt[fac.getFName()][day][time] == '' and '5' in sec:
#     fac_day = faculty_tt[fac.getFName()][day]
#     if time == 0:
#         if fac_day[time + 1] == fac_day[time + 2] == fac_day[time + 3] == '':
#             pass
#         elif fac_day[1] == '' and (fac_day[2].__contains__('/') or fac_day[2].__contains__('LAB')):
#             pass
#         elif (fac_day[time + 1] == fac_day[time + 3] == '') and fac_day[time + 2] != '':
#             pass
#         elif fac_day[time + 1] != '' and fac_day[time + 2] == '' and fac_day[time + 3] != '':
#             pass
#         elif fac_day[time + 1] == '' and fac_day[time + 2] != '' and fac_day[time + 3] != '':
#             pass
#         else:
#             return False
#
#     elif time == 1:
#         if fac_day[time - 1] == '' and fac_day[time + 1] == '':
#             pass
#         elif fac_day[time + 1].__contains__('LAB') and (fac_day[time + 1][0:5] in section_labs[sec].keys()):
#             pass
#         # elif fac_day[time+1].__contains__('/') and (fac_day[time + 1] in section_labs[sec].keys())
#         elif fac_day[time - 1] == sub and fac_day[time + 1] == '':
#             pass
#
#         elif fac_day[time - 1] == '' and fac_day[time + 1] == sub:
#             pass
#         else:
#             return False
#
#     elif time == 2:
#         if fac_day[time - 1] == '' and fac_day[time + 1] == '':
#             pass
#         elif fac_day[time + 1][0:5] in section_labs[sec].keys():
#             pass
#         elif fac_day[time - 1] == '' and fac_day[time - 2] != sub:
#             pass
#         elif fac_day[time - 1] == sub and fac_day[time + 1] == '':
#             pass
#         elif fac_day[time - 1] == '' and fac_day[time + 1] == sub:
#             pass
#         else:
#             return False
#
#     elif time == 3:
#         if fac_day[time - 1] == '':
#             pass
#         elif fac_day[time + 1][0:5] in section_labs[sec].keys():
#             pass
#         elif fac_day[time - 2] != sub:
#             pass
#         elif fac_day[time - 1] == sub:
#             pass
#         else:
#             return False
#
#     elif time == 5:
#         if fac_day[time + 1] == '':
#             pass
#         elif fac_day[time + 1][0:5] in section_labs[sec].keys():
#             pass
#         elif fac_day[time + 1] == sub and fac_day[time + 2] == '':
#             pass
#         # elif fac_day[time + 1] == sub:
#         #     pass
#         elif fac_day[time + 1] == '' and fac_day[time + 2] != '':
#             pass
#         else:
#             return False
#
#     elif time == 6:
#         if fac_day[time - 1] == '' and fac_day[time + 1] == '':
#             pass
#         elif fac_day[time + 1] == '' and fac_day[time - 1] == sub:
#             pass
#         elif fac_day[time + 1] == sub and fac_day[time - 1] == '':
#             pass
#         else:
#             return True
#
#     else:
#         if fac_day[time - 1] == '' and fac_day[time - 2] == '':
#             pass
#         elif fac_day[time - 1] == sub and fac_day[time - 2] == '':
#             pass
#         # elif fac_day[time - 1] == sub:
#         #     pass
#         elif fac_day[time - 1] == '' and fac_day[time - 2] != '':
#             pass
#         else:
#             return False
#     pass
# else:
#     return False
def available_theory(sec, day, time, sub):
    global section_tt, faculty_tt, section_th_fac

    th_fac = section_th_fac[sec]

    # sec1 = tt[day][0:4]
    # sec2 = tt[day][5:]

    if section_tt[sec][day][time] == '':
        pass
    else:
        return False

    # if sub in section_tt[sec][day]:
    #     return False
    if section_tt[sec][day].count(sub) < 2:
        pass
    else:
        return False

    if sub == 'M/L':
        fac = mentors[sec]
    else:
        fac = th_fac[sub]

    fac_day = faculty_tt[fac.getFName()][day]
    if fac_day[time] == '':
        pass
    else:
        return False

    if time == 0:
        if fac_day[1] == fac_day[2] == fac_day[3] == '':
            pass
        elif (fac_day[1] == '') and (fac_day[2].__contains__('/') or fac_day[2].__contains__('LAB')):
            pass
        elif (fac_day[time + 1] == fac_day[time + 3] == '') and (fac_day[time + 2] != ''):
            pass
        elif (fac_day[time + 1] != '') and (fac_day[time + 2] == '') and (fac_day[time + 3] != ''):
            pass
        elif (fac_day[time + 1] == '') and (fac_day[time + 2] != '') and (fac_day[time + 3] != ''):
            pass
        else:
            return False

    elif time == 1:
        if fac_day[0] == fac_day[2] == fac_day[3] == '':
            pass
        elif (fac_day[time - 1] == fac_day[time + 2] == '') and (fac_day[time + 1] != ''):
            pass
        elif fac_day[0] != '' and fac_day[2] == '' and fac_day[3] == '':
            pass
        elif fac_day[0] != '' and fac_day[3] != '' and fac_day[2] == '':
            pass
        elif fac_day[0] == '' and (fac_day[2].__contains__('/') or fac_day[2].__contains__('LAB')):
            pass
        else:
            return False

    elif time == 2:
        if fac_day[0] == fac_day[1] == fac_day[3] == '':
            pass
        elif (fac_day[0].__contains__('/') or fac_day[0].__contains__('LAB')) and fac_day[3] == '':
            pass
        elif fac_day[0] != '' and fac_day[1] == '' and fac_day[3] != '':
            pass
        elif fac_day[0] == fac_day[1] == '' and fac_day[3] != '':
            pass
        else:
            return False

    elif time == 3:
        if fac_day[0] == fac_day[1] == fac_day[2] == '':
            pass
        elif (fac_day[0].__contains__('/') or fac_day[0].__contains__('LAB')) and fac_day[2] == '':
            pass
        elif fac_day[0] == fac_day[1] == '' and fac_day[2] != '':
            pass
        elif fac_day[0] != '' and fac_day[1] != '' and fac_day[2] == '':
            pass
        elif fac_day[0] == '' and fac_day[2] == '' and fac_day[1] != '':
            pass
        elif fac_day[0] != '' and fac_day[2] != '' and fac_day[1] == '':
            pass
        else:
            return False

    elif time == 5:
        if fac_day[6] == '' and fac_day[7] == '':
            pass
        elif fac_day[6] == '' and fac_day[7] != '':
            pass
        elif fac_day[6] != '' and fac_day[7] == '':
            pass
        elif fac_day[6].__contains__('/') or fac_day[6].__contains__('LAB'):
            pass
        else:
            return False

    elif time == 6:
        if fac_day[5] == fac_day[7] == '':
            pass
        elif fac_day[5] == '' and fac_day[7] != '':
            pass
        elif fac_day[5] != '' and fac_day[7] == '':
            pass
        else:
            return False

    else:
        if fac_day[5] == fac_day[6] == '':
            pass
        elif fac_day[5] == '' and fac_day[6] != '':
            pass
        elif fac_day[5] != '' and fac_day[6] == '':
            pass
        else:
            return False

    # if sub != 'M/L':
    #     load = fac.getFLoad()[sec]
    #     for i in load:
    #         if i[0] == sub:
    #             if 6 >= i[2] >= 1:
    #                 break
    #             else:
    #                 return False

    return True


# slots=[]
# sem_5_sec.reverse()
# sem_3_sec.reverse()
# sec_slots=[]
last_sec1 = ''
last_sec2 = ''


def set_theory():
    global section_tt, faculty_tt, last_sec1, last_sec2

    for sec in sem_5_sec:
        print(sec)
        # tt = section_tt[sec]
        # for i in tt:
        #     print(i)
        # if sec=='AIML 3C':
        #     continue
        if visited[sec]:
            continue
        # if ('CSE 3A' in sec) or ('CY 3A' in sec) or ('AIML 3A' in sec):
        #     sec_slots=[]
        c = 0
        while c < 1:
            day = random.choice([0, 1, 2, 4, 3])
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            fac = mentors[sec]
            if available_theory(sec, day, time, 'M/L'):
                lib_slots[sec] = [day, time, sec, fac]
                section_tt[sec][day][time] = 'M/L'
                section_tt['Lib'][day][time] = sec
                faculty_tt[fac.getFName()][day][time] = sec + '\n' + 'M/L'
                c = c + 1
                # print('M/L')
                break

        slot_list = []
        th_fac = section_th_fac[sec]
        unalloted = {}
        c = 0
        sub_list = []
        sub_list = list(th_fac.keys())
        # print(sub_list)
        sub_count = {}
        if '5' in sec:
            for sub in sub_list:
                sub_count[sub] = 6
        # if '3' in sec:
        #     for sub in sub_list:
        #         sub_count[sub] = 5
        sub_list = list(th_fac.keys())
        while checkEmpty(sec):

            c = c + 1
            # for day in [0, 1, 2, 3, 4]:
            #     for time in [0, 1, 2, 3, 5, 6, 7]:
            day = random.choice([0, 1, 2, 3, 4])
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            if checkEmpty(sec):
                visited[sec] = True
                pass
            else:
                break
            sub = random.choice(sub_list)
            if sub not in unalloted.keys():
                unalloted[sub] = 0
            if available_theory(sec, day, time, sub):

                fac = th_fac[sub]
                slot_list.append([day, time, fac])
                section_tt[sec][day][time] = sub
                faculty_tt[fac.getFName()][day][time] = sub
                # print(sub)
                sub_count[sub] = sub_count[sub] - 1
                if sub_count[sub] == 0:
                    sub_list.remove(sub)
                del unalloted[sub]


            else:
                unalloted[sub] = unalloted[sub] + 1

                if unalloted[sub] > 100:
                    c = 0
                    while len(slot_list) != 0:
                        i = slot_list[0]
                        section_tt[sec][i[0]][i[1]] = ''
                        faculty_tt[i[2].getFName()][i[0]][i[1]] = ''
                        slot_list.remove(i)
                        c = c + 1

                    slot = lib_slots[sec]
                    day = slot[0]
                    time = slot[1]
                    fac = slot[3]

                    section_tt[sec][day][time] = ''
                    section_tt['Lib'][day][time] = ''
                    faculty_tt[fac.getFName()][day][time] = ''

                    visited[sec] = False

                    # print("removed", c)
                    last_sec2 = last_sec1
                    last_sec1 = sec
                    return False

    for sec in sem_3_sec:
        print(sec)
        if sec == 'AIML 3C':
            # tt=section_tt[sec]
            # for i in tt:
            #     print(i)
            continue
        if visited[sec]:
            continue
        c = 0

        # if ('CSE 3A' in sec) or ('CY 3A' in sec) or ('AIML 3A' in sec):
        #     sec_slots=[]

        while c < 1:
            day = random.choice([0, 1, 2, 4, 3])
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            fac = mentors[sec]
            if available_theory(sec, day, time, 'M/L'):
                lib_slots[sec] = [day, time, sec, fac]
                section_tt[sec][day][time] = 'M/L'
                section_tt['Lib'][day][time] = sec
                faculty_tt[fac.getFName()][day][time] = sec + '\n' + 'M/L'
                c = c + 1
                # print('M/L')
                break

        slot_list = []
        th_fac = section_th_fac[sec]
        unalloted = {}
        # c = 0
        sub_list = list(th_fac.keys())
        # print(sub_list)
        sub_count = {}
        for sub in sub_list:
            # if '01' in sub:
            #     sub_count[sub]=4
            # else:
            sub_count[sub] = 5
        while checkEmpty(sec):
            # sub_list = list(th_fac.keys())
            # c = c + 1
            # for day in [0, 1, 2, 3, 4]:
            #     for time in [0, 1, 2, 3, 5, 6, 7]:
            day = random.choice([0, 1, 2, 3, 4])
            time = random.choice([0, 1, 2, 3, 5, 6, 7])
            if checkEmpty(sec):
                visited[sec] = True
                pass
            else:
                break
            sub = random.choice(sub_list)
            if sub not in unalloted.keys():
                unalloted[sub] = 0
            if available_theory(sec, day, time, sub):
                # print(sub)
                fac = th_fac[sub]
                slot_list.append([day, time, fac])
                section_tt[sec][day][time] = sub
                faculty_tt[fac.getFName()][day][time] = sub
                # for i in range(len(fac.getFLoad()[sec])):
                #     if fac.getFLoad()[sec][i][0] == sub:
                #         fac.getFLoad()[sec][i][2] = fac.getFLoad()[sec][i][2] - 1
                # print("pass")
                sub_count[sub] = sub_count[sub] - 1
                if sub_count[sub] == 0:
                    sub_list.remove(sub)
                del unalloted[sub]

            else:
                unalloted[sub] = unalloted[sub] + 1

                if unalloted[sub] > 100:
                    c = 0
                    while len(slot_list) != 0:
                        i = slot_list[0]
                        section_tt[sec][i[0]][i[1]] = ''
                        faculty_tt[i[2].getFName()][i[0]][i[1]] = ''
                        slot_list.remove(i)
                        c = c + 1

                    slot = lib_slots[sec]
                    day = slot[0]
                    time = slot[1]
                    fac = slot[3]

                    section_tt[sec][day][time] = ''
                    section_tt['Lib'][day][time] = ''
                    faculty_tt[fac.getFName()][day][time] = ''
                    visited[sec] = False
                    # print("removed", c)
                    last_sec2 = last_sec1
                    last_sec1 = sec
                    return False

    # slots=[]

    return True


def setTimeTables():
    # reset_tt()
    print("Setting labs")
    while True:
        c2 = 0
        c = 0

        comp = 0
        # print("Setting labs")
        while True:
            # if c > 0:
            #     pass
            # reset_tt()
            # reset_load()
            # reset_visited()
            # set_load()
            c = c + 1

            if set_labs():
                print('Labs Set1')
                reset_visited()
                comp = 1
                break

            # if last_sec.__contains__('AIML')or last_sec.__contains__('CY'):
            # if last_sec.__contains__('AIML 3C'):
            #     if c > 300:
            #         reset_tt()
            #         reset_visited()
            #         c = 0
            #         # for i,j in lab_counter.items():
            #         #     print(i,j)
            #         # comp=1
            #         break
            #     else:
            #         continue
            if c > 100:
                # copy_prev_data()
                reset_presets()
                reset_tt()
                reset_visited()
                reset_lab_day_counter()
                # reset_lab_counter()
                # reset_load()
                # set_load()
                c = 0
                break

            # else:
            #     continue

            # reset_load()
            # set_load()
        c2 = 0
        # for sec in sec_list:
        #     print(sec)
        #     tt=section_tt[sec]
        #     for i in tt:
        #         print(i)
        # t=2000
        if comp == 1:
            while True:
                val = 0
                c2 = c2 + 1

                # if set_lib():
                #     print("Lib set")
                #     pass
                # else:
                #     continue
                if set_theory():
                    comp = 2
                    break

                # reset_load()
                # set_load()
                if last_sec1 == last_sec2:
                    pass
                else:
                    c2 = 0
                if c2 > 15000:
                    reset_visited()
                    # reset_load()
                    # set_load()
                    break
        if comp == 2:
            break


# section_tt={
#     'CSE 3A':[['CS303', 'CS301', 'CS303/CS304', 'CS303/CS304', 'Lunch', 'CS304', 'CS305/CS306', 'CS305/CS306'],
# ['CS303/CS304', 'CS303/CS304', 'CS303', 'CS301', 'Lunch', 'CS303', 'CS302', 'CS304'],
# ['CS305/CS306', 'CS305/CS306', 'CS305', 'CS304', 'Lunch', 'CS301', 'CS302', 'CS305'],
# ['CS304', 'CS304', 'CS305', 'CS303', 'Lunch', 'CS303', 'CS302', 'CS301'],
# ['CS302', 'M/L', 'CS307LAB', 'CS307LAB', 'Lunch', 'CS305', 'CS305', 'CS302']],
#     'CSE 3B':[['CS304', 'CS303', 'CS304', 'CS301', 'Lunch', 'CS301', 'CS303', 'CS305'],
# ['CS302', 'CS302', 'CS305', 'CS303', 'Lunch', 'CS303', 'CS307LAB', 'CS307LAB'],
# ['CS303/CS304', 'CS303/CS304', 'CS305', 'CS304', 'Lunch', 'CS301', 'CS303/CS304', 'CS303/CS304'],
# ['M/L', 'CS301', 'CS301', 'CS303', 'Lunch', 'CS302', 'CS305/CS306', 'CS305/CS306'],
# ['CS304', 'CS302', 'CS305/CS306', 'CS305/CS306', 'Lunch', 'CS304', 'CS305', 'CS305']],
#     'CSE 3C':[['CS305', 'CS302', 'CS305/CS306', 'CS305/CS306', 'Lunch', 'CS301', 'CS303', 'CS301'],
# ['CS304', 'CS302', 'CS305/CS306', 'CS305/CS306', 'Lunch', 'CS302', 'CS304', 'CS305'],
# ['CS301', 'CS301', 'CS303', 'CS302', 'Lunch', 'CS305', 'CS304', 'CS303'],
# ['CS304', 'CS301', 'CS303/CS304', 'CS303/CS304', 'Lunch', 'CS302', 'CS303/CS304', 'CS303/CS304'],
# ['CS305', 'CS303', 'M/L', 'CS303', 'Lunch', 'CS305', 'CS307LAB', 'CS307LAB']],
#     'CSE 5A':[['CS501', 'CS504', 'CS502LAB', 'CS502LAB', 'Lunch', 'CS504', 'CS503', 'CS501'],
# ['CS503', 'CS501', 'CS503', 'CS504', 'Lunch', 'CS504', 'CS502', 'CS502'],
# ['CS503', 'CS502', 'M/L', 'CS504', 'Lunch', 'CS504', 'CS502', 'CS501'],
# ['CS503', 'CS504', 'CS508LAB', 'CS508LAB', 'Lunch', 'CS501', 'CS505LAB', 'CS505LAB'],
# ['CS504', 'CS503', 'CS506LAB', 'CS506LAB', 'Lunch', 'CS502', 'CS501LAB', 'CS501LAB']],
#     'CSE 5C':[['CS501', 'CS504', 'CS508LAB', 'CS508LAB', 'Lunch', 'CS504', 'CS503', 'CS501'],
# ['CS501', 'CS503', 'CS501LAB', 'CS501LAB', 'Lunch', 'CS503', 'CS502', 'CS501'],
# ['CS502', 'CS502', 'CS506LAB', 'CS506LAB', 'Lunch', 'M/L', 'CS505LAB', 'CS505LAB'],
# ['CS502LAB', 'CS502LAB', 'CS502', 'CS501', 'Lunch', 'CS501', 'CS504', 'CS504'],
# ['CS502', 'CS501', 'CS504', 'CS503', 'Lunch', 'CS504', 'CS502', 'CS503']],
#     'CSE 5B':[['CS501', 'M/L', 'CS504', 'CS503', 'Lunch', 'CS502', 'CS506LAB', 'CS506LAB'],
# ['CS502LAB', 'CS502LAB', 'CS504', 'CS502', 'Lunch', 'CS501', 'CS503', 'CS502'],
# ['CS501LAB', 'CS501LAB', 'CS502', 'CS503', 'Lunch', 'CS503', 'CS504', 'CS501'],
# ['CS503', 'CS503', 'CS502', 'CS502', 'Lunch', 'CS501', 'CS508LAB', 'CS508LAB'],
# ['CS504', 'CS501', 'CS501', 'CS503', 'Lunch', 'CS504', 'CS505LAB', 'CS505LAB']],
#     'CY 3A':[['M/L', 'CY304', 'CY301', 'CY301', 'Lunch', 'CY305', 'CY303', 'CY303'],
# ['CY305', 'CY301', 'CY304', 'CY302', 'Lunch', 'CY304', 'CY305/CY306', 'CY305/CY306'],
# ['CY305', 'CY302', 'CY307LAB', 'CY307LAB', 'Lunch', 'CY305', 'CY302', 'CY304'],
# ['CY303/CY304', 'CY303/CY304', 'CY305/CY306', 'CY305/CY306', 'Lunch', 'CY303', 'CY302', 'CY303'],
# ['CY304', 'CY302', 'CY303/CY304', 'CY303/CY304', 'Lunch', 'CY301', 'CY303', 'CY301']],
#     'CY 3B':[['CY301', 'CY305', 'CY303', 'CY303', 'Lunch', 'CY305', 'CY303/CY304', 'CY303/CY304'],
# ['CY301', 'M/L', 'CY303', 'CY303', 'Lunch', 'CY302', 'CY303/CY304', 'CY303/CY304'],
# ['CY304', 'CY303', 'CY302', 'CY305', 'Lunch', 'CY301', 'CY305/CY306', 'CY305/CY306'],
# ['CY307LAB', 'CY307LAB', 'CY304', 'CY305', 'Lunch', 'CY304', 'CY302', 'CY305'],
# ['CY304', 'CY302', 'CY301', 'CY301', 'Lunch', 'CY304', 'CY305/CY306', 'CY305/CY306']],
#     'CY 5A':[['CY501LAB', 'CY501LAB', 'CY505LAB', 'CY505LAB', 'Lunch', 'CY502', 'CY504', 'CY501'],
# ['CY502', 'CY502', 'CY506LAB', 'CY506LAB', 'Lunch', 'CY504', 'CY501', 'CY501'],
# ['CY508LAB', 'CY508LAB', 'CY503', 'CY501', 'Lunch', 'CY504', 'CY502', 'M/L'],
# ['CY502', 'CY501', 'CY502LAB', 'CY502LAB', 'Lunch', 'CY501', 'CY502', 'CY504'],
# ['CY503', 'CY501', 'CY503', 'CY501', 'Lunch', 'CY502', 'CY502', 'CY504']],
#     'CY 5B':[['CY504', 'CY503', 'CY504', 'CY502', 'Lunch', 'CY501', 'CY508LAB', 'CY508LAB'],
# ['CY501LAB', 'CY501LAB', 'CY504', 'CY502', 'Lunch', 'CY503', 'CY506LAB', 'CY506LAB'],
# ['CY504', 'CY502', 'CY502LAB', 'CY502LAB', 'Lunch', 'CY504', 'CY503', 'CY503'],
# ['CY502', 'CY503', 'CY505LAB', 'CY505LAB', 'Lunch', 'CY501', 'CY504', 'CY503'],
# ['M/L', 'CY504', 'CY502', 'CY502', 'Lunch', 'CY501', 'CY501', 'CY504']],
#     'AIML 3A':[['AL305/AL306', 'AL305/AL306', 'AL303/AL304', 'AL303/AL304', 'Lunch', 'AL305', 'AL303', 'AL304'],
# ['AL302', 'AL303', 'AL305', 'AL303', 'Lunch', 'AL305', 'AL301', 'AL304'],
# ['AL303/AL304', 'AL303/AL304', 'AL305/AL306', 'AL305/AL306', 'Lunch', 'AL302', 'AL304', 'AL305'],
# ['AL303', 'AL304', 'AL305', 'AL301', 'Lunch', 'AL302', 'AL303', 'AL301'],
# ['AL307LAB', 'AL307LAB', 'AL304', 'AL302', 'Lunch', 'AL302', 'M/L', 'AL301']],
#     'AIML 3B':[['AL303', 'AL302', 'AL305', 'AL304', 'Lunch', 'AL301', 'AL302', 'AL301'],
# ['AL307LAB', 'AL307LAB', 'AL303/AL304', 'AL303/AL304', 'Lunch', 'AL301', 'AL302', 'M/L'],
# ['AL302', 'AL301', 'AL304', 'AL304', 'Lunch', 'AL301', 'AL302', 'AL303'],
# ['AL305/AL306', 'AL305/AL306', 'AL305', 'AL304', 'Lunch', 'AL303', 'AL303', 'AL304'],
# ['AL305/AL306', 'AL305/AL306', 'AL303/AL304', 'AL303/AL304', 'Lunch', 'AL303', 'AL305', 'AL305']],
#     'AIML 3C':[['AL303', 'AL302', 'AL304', 'AL302', 'Lunch', 'M/L', 'AL303/AL304', 'AL303/AL304'],
# ['AL303', 'AL301', 'AL304', 'AL305', 'Lunch', 'AL305', 'AL301', 'AL304'],
# ['AL302', 'AL301', 'AL307LAB', 'AL307LAB', 'Lunch', 'AL305', 'AL302', 'AL303'],
# ['AL305/AL306', 'AL305/AL306', 'AL303', 'AL304', 'Lunch', 'AL301', 'AL305/AL306', 'AL305/AL306'],
# ['AL305', 'AL305', 'AL303', 'AL301', 'Lunch', 'AL304', 'AL303/AL304', 'AL303/AL304']],
#     'AIML 5A':[['AL502LAB', 'AL502LAB', 'AL503', 'AL503', 'Lunch', 'AL502', 'AL508LAB', 'AL508LAB'],
# ['M/L', 'AL503', 'AL504', 'AL501', 'Lunch', 'AL504', 'AL505LAB', 'AL505LAB'],
# ['AL503', 'AL502', 'AL503', 'AL501', 'Lunch', 'AL504', 'AL501', 'AL504'],
# ['AL502', 'AL504', 'AL503', 'AL504', 'Lunch', 'AL501', 'AL502', 'AL501'],
# ['AL501LAB', 'AL501LAB', 'AL506LAB', 'AL506LAB', 'Lunch', 'AL501', 'AL502', 'AL502']],
#     'AIML 5B':[['AL506LAB', 'AL506LAB', 'AL502', 'M/L', 'Lunch', 'AL503', 'AL502', 'AL501'],
# ['AL501', 'AL503', 'AL501', 'AL503', 'Lunch', 'AL504', 'AL502', 'AL502'],
# ['AL504', 'AL501', 'AL502', 'AL503', 'Lunch', 'AL504', 'AL501LAB', 'AL501LAB'],
# ['AL504', 'AL503', 'AL508LAB', 'AL508LAB', 'Lunch', 'AL503', 'AL505LAB', 'AL505LAB'],
# ['AL502LAB', 'AL502LAB', 'AL504', 'AL503', 'Lunch', 'AL501', 'AL502', 'AL501']],
#     'AIML 5C':[['AL508LAB', 'AL508LAB', 'AL503', 'AL502', 'Lunch', 'AL503', 'AL504', 'AL501'],
# ['AL502LAB', 'AL502LAB', 'AL501', 'AL504', 'Lunch', 'AL503', 'AL501LAB', 'AL501LAB'],
# ['AL504', 'AL502', 'AL505LAB', 'AL505LAB', 'Lunch', 'AL504', 'AL506LAB', 'AL506LAB'],
# ['AL502', 'AL501', 'M/L', 'AL501', 'Lunch', 'AL502', 'AL504', 'AL503'],
# ['AL501', 'AL504', 'AL502', 'AL501', 'Lunch', 'AL503', 'AL504', 'AL503']]
# }

fac_info_row = ['14', '15', '16', '17', '18', '19', '20', '21']


def copy_tt_to_excel():
    for sec in sec_list:
        print(sec)
        tt = section_tt[sec]
        for i in tt:
            print(i)
        sheet = section_sheets[sec]
        labs = section_labs[sec]
        labs_fac = section_labs_fac[sec]
        th_fac = section_th_fac[sec]
        for day in [0, 1, 2, 3, 4]:
            for time in [0, 1, 2, 3, 5, 6, 7]:
                if tt[day][time] == '':
                    continue

                elif tt[day][time] == 'NA' or tt[day][time] == 'Na':
                    sheet[time_cell[time] + day_cell[day]] = tt[day][time]
                    continue

                elif tt[day][time] == 'M/L':
                    sheet[time_cell[time] + day_cell[day]] = 'M/L'
                    section_sheets['Lib'][time_cell[time] + day_cell[day]] = sec
                    faculty_sheets[mentors[sec].getFName()][time_cell[time] + day_cell[day]] = sec + '\n' + 'M/L'
                    continue

                elif tt[day][time] == sheet[time_cell[time] + day_cell[day]].value:
                    continue

                elif len(tt[day][time]) == 5:
                    sub = tt[day][time]
                    fac = th_fac[sub]
                    sheet[time_cell[time] + day_cell[day]] = sub
                    # print(fac.getFName())
                    # print(sub)
                    # print(faculty_sheets[fac.getFName()][time_cell[time-1] + day_cell[day]].value)
                    # print('Time',time)
                    faculty_sheets[fac.getFName()][time_cell[time] + day_cell[day]] = sec + "\n" + sub
                    continue

                elif tt[day][time] != '' and sheet[time_cell[time] + day_cell[day]].value == None:
                    if ('/' in tt[day][time]) and (time not in [1, 3, 7]):
                        lab = tt[day][time][0:11]
                        sheet.merge_cells(time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        sheet[time_cell[time] + day_cell[day]] = lab + "\n" + \
                                                                 labs[lab][0] + "/" + labs[lab][1]
                        if '8' in lab:
                            sheet['E' + fac_info_row[int(lab[4]) - 2]] = labs[lab][0] + ' Building LNCT&S'
                            sheet['E' + fac_info_row[int(lab[10]) - 2]] = labs[lab][1] + ' Building LNCT&S'
                        else:
                            sheet['E' + fac_info_row[int(lab[4]) - 1]] = labs[lab][0] + ' Building LNCT&S'
                            sheet['E' + fac_info_row[int(lab[10]) - 1]] = labs[lab][1] + ' Building LNCT&S'
                        lab_sheets[labs[lab][0]][time_cell[time] + day_cell[day]] = \
                            lab_sheets[labs[lab][1]][time_cell[time] + day_cell[day]] = sec
                        faculty_sheets[labs_fac[lab][0].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][1].getFName()].merge_cells(
                            time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                        faculty_sheets[labs_fac[lab][0].getFName()][time_cell[time] + day_cell[day]] = \
                            faculty_sheets[labs_fac[lab][1].getFName()][
                                time_cell[time] + day_cell[day]] = sec + "\n" + lab + "\n" + labs[lab][0] + '/' + \
                                                                   labs[lab][1]

                        continue
                    elif ('LAB' in tt[day][time]) and (time not in [1, 3, 7]):
                        if tt[day][time][0:5] in section_labs[sec].keys():
                            if time in [0, 2, 6]:
                                pass
                            else:
                                continue
                            labs_fac = section_labs_fac[sec]
                            lab = tt[day][time][0:5]
                            sheet.merge_cells(
                                time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                            sheet[time_cell[time] + day_cell[day]] = lab + "\n" + labs[lab][0]
                            if '8' in lab:
                                sheet['E' + fac_info_row[int(lab[4]) - 2]] = labs[lab][0] + ' Building LNCT&S'
                                # sheet['E' + fac_info_row[int(lab[10]) - 2]] = labs[lab][1] + ' Building LNCT&S'
                            else:
                                sheet['E' + fac_info_row[int(lab[4]) - 1]] = labs[lab][0] + ' Building LNCT&S'
                                # sheet['E' + fac_info_row[int(lab[10]) - 1]] = labs[lab][1] + ' Building LNCT&S'
                            lab_sheets[labs[lab][0]][time_cell[time] + day_cell[day]] = sec
                            faculty_sheets[labs_fac[lab][0].getFName()].merge_cells(
                                time_cell[time] + day_cell[day] + ":" + time_cell[time + 1] + day_cell[day])
                            faculty_sheets[labs_fac[lab][0].getFName()][
                                time_cell[time] + day_cell[day]] = sec + "\n" + lab + '\n' + labs[lab][0]
                            continue
                    # elif (('LAB' in tt[day][time]) or ('/' in tt[day][time])) and (time in [1,3,7]):
                    #     pass
                    # else:
                    #     sub = tt[day][time]
                    #     # if (('/' in sub) or ('LAB' in sub)) and time in [1,3,7]:
                    #     #     continue
                    #     # else:
                    #     #     pass
                    #
                    #     fac = th_fac[sub]
                    #
                    #     # if isinstance()
                    #
                    #     # cell=faculty_sheets[fac.getFName()][time_cell[time] + day_cell[day]]
                    #     # if isinstance(cell,MergedCell):
                    #     #     continue
                    #     sheet[time_cell[time] + day_cell[day]] = sub
                    #     print(fac.getFName())
                    #     print(sub)
                    #     print(faculty_tt[fac.getFName()][day][time-1])
                    #     print(faculty_tt[fac.getFName()][day][time])
                    #     faculty_sheets[fac.getFName()][time_cell[time] + day_cell[day]] = sec + "\n" + sub


def set_sheets():
    files = next(os.walk('C:/Users/Swadesh Sharma/Desktop/SchedulXpert/'))[1]
    if not files:
        n = 0
    else:
        n = len(files)
    newpath = 'C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test ' + str(n + 1)

    # Creating new directories
    if not os.path.exists(newpath):
        os.makedirs(newpath)
        os.makedirs(newpath + '/Faculty')
        os.makedirs(newpath + '/Sections')
        os.makedirs(newpath + '/Labs')

    # saving the files in the directory creates
    for i in workbooks.keys():
        if 'Lib' in i or 'CS' in i or 'AIML' in i or 'CY' in i:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Sections/" + i + ".xlsx")
        elif 'Dr' in i or 'Prof' in i:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Faculty/" + i + ".xlsx")
        else:
            workbooks[i].save(
                "C:/Users/Swadesh Sharma/Desktop/SchedulXpert/Test " + str(n + 1) + "/Labs/" + i + ".xlsx")

# print(sec_list)
# print('Started')
# setTimeTables()
# for sec in sec_list:
#     print(sec)
#     tt = section_tt[sec]
#     for i in tt:
#         print(i)
# copy_tt_to_excel()
# set_sheets()
# print('finished')

# keep track of the slots and don't repeat the same slot for same lab
